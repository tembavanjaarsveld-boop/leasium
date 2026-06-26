"""Mark active tenants from the SKJ portfolio workbook as already enrolled.

The portfolio import (`scripts/import_portfolio_workbook.py`) already loaded
properties, units, tenants, and leases from the SKJ source-of-truth workbook.
But the tenants in that workbook are real, currently-occupying tenants who
were enrolled long before Relby existed — they don't need to go through
the onboarding workflow. This script tags them as already complete:

  - Fills in missing tenant record fields (trading name, contact email,
    notes) from the spreadsheet. Does NOT overwrite existing values.
  - Creates or updates a `tenant_onboarding` record per tenant set to
    status=applied, with submitted_at / reviewed_at / applied_at = now
    and applied_by = the operator running the script. The token is a
    deterministic synthetic value (this enrollment never used a portal
    link).
  - Records source provenance in `tenant_metadata.enrollment_source`
    and `tenant_onboarding.review_data.enrollment_source` so the script
    is idempotent and re-runnable.

Provider-mutation guardrail (CLAUDE.md §2.1): this script does NOT send
any email, SMS, or portal invite. It does not call SendGrid / Twilio /
Xero. It is pure SQL writes to the local registers via SQLAlchemy.

Default is dry-run. Pass --apply to actually mutate. Pass --workbook to
override the source workbook path.

Usage on the Mac:

    cd ~/Documents/Stewart
    .venv/bin/python -m scripts.enroll_from_workbook \\
        --workbook "/path/to/SKJ_Property_Portfolio_Source_of_Truth.xlsx" \\
        --entity-id <ENTITY_UUID> \\
        --operator-email temba@skjcapital.com

That prints a plan. Once it looks right, add --apply:

    .venv/bin/python -m scripts.enroll_from_workbook \\
        --workbook "/path/to/SKJ_Property_Portfolio_Source_of_Truth.xlsx" \\
        --entity-id <ENTITY_UUID> \\
        --operator-email temba@skjcapital.com \\
        --apply
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker
from stewart.core.models import (
    AppUser,
    Entity,
    Lease,
    Property,
    TenancyUnit,
    Tenant,
    TenantOnboarding,
    TenantOnboardingStatus,
)
from stewart.core.settings import Settings

# Idempotency marker. Re-running the script with the same workbook sees
# this in tenant_metadata.enrollment_source and skips the tenant.
ENROLLMENT_SOURCE = "skj_portfolio_already_enrolled_2026_05_23"

EMAIL_RE = re.compile(r"^[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}$")

# Status values in the workbook that count as "currently enrolled". Anything
# else (Vacant, Pending, Drafting, blanks) is skipped.
ENROLLED_STATUSES = {"active", "holdover", "holding over", "month-to-month"}


@dataclass
class TenancyRow:
    sheet_row: int
    property_code: str
    unit_code: str
    legal_name: str
    trading_name: str | None
    primary_contact: str | None
    notes: str | None
    status: str


def _clean(value: Any) -> str | None:
    """Strip a workbook cell down to a non-empty string, or None.

    Treats '—' / '-' / 'Not specified' as null (the workbook uses these as
    explicit "no value" markers in places).
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text in {"—", "-", "Not specified", "N/A", "TBC", "TBD"}:
        return None
    return text


def _is_email(value: str | None) -> bool:
    return bool(value and EMAIL_RE.match(value.strip()))


def _read_tenancies(workbook_path: Path) -> list[TenancyRow]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "Tenancies" not in wb.sheetnames:
        raise SystemExit(
            f"Workbook {workbook_path} has no 'Tenancies' sheet."
        )
    ws = wb["Tenancies"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]

    def col(name: str) -> int | None:
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx_property = col("Property")
    idx_unit = col("Unit code")
    idx_legal = col("Tenant (legal name)")
    idx_trading = col("Trading name")
    idx_contact = col("Primary contact")
    idx_notes = col("Notes")
    idx_status = col("Status")

    required = {
        "Property": idx_property,
        "Unit code": idx_unit,
        "Tenant (legal name)": idx_legal,
        "Status": idx_status,
    }
    missing = [name for name, i in required.items() if i is None]
    if missing:
        raise SystemExit(
            f"Workbook is missing required columns in Tenancies sheet: {missing}"
        )

    out: list[TenancyRow] = []
    for row_idx, raw in enumerate(rows[1:], start=2):
        property_code = _clean(raw[idx_property])
        unit_code = _clean(raw[idx_unit])
        legal = _clean(raw[idx_legal])
        status = _clean(raw[idx_status]) or ""
        if not (property_code and unit_code and legal):
            continue
        if legal.upper() == "VACANT":
            continue
        if status.lower() not in ENROLLED_STATUSES:
            continue
        out.append(
            TenancyRow(
                sheet_row=row_idx,
                property_code=property_code,
                unit_code=unit_code,
                legal_name=legal,
                trading_name=_clean(raw[idx_trading]) if idx_trading is not None else None,
                primary_contact=_clean(raw[idx_contact]) if idx_contact is not None else None,
                notes=_clean(raw[idx_notes]) if idx_notes is not None else None,
                status=status,
            )
        )
    return out


@dataclass
class MatchResult:
    row: TenancyRow
    property_id: UUID | None = None
    tenancy_unit_id: UUID | None = None
    tenant: Tenant | None = None
    lease: Lease | None = None
    onboarding: TenantOnboarding | None = None
    blockers: list[str] | None = None

    def blocked(self) -> bool:
        return bool(self.blockers)


def _match_row(session: Session, entity_id: UUID, row: TenancyRow) -> MatchResult:
    """Resolve a workbook row to existing register records."""
    result = MatchResult(row=row, blockers=[])

    # Property: portfolio_code in metadata.
    prop = next(
        (
            p
            for p in session.scalars(
                select(Property).where(
                    Property.entity_id == entity_id,
                    Property.deleted_at.is_(None),
                )
            )
            if (p.property_metadata or {}).get("portfolio_code") == row.property_code
        ),
        None,
    )
    if prop is None:
        result.blockers.append(f"No property with portfolio_code={row.property_code!r}")
        return result
    result.property_id = prop.id

    # Tenancy unit by property + label.
    unit = session.scalars(
        select(TenancyUnit).where(
            TenancyUnit.property_id == prop.id,
            TenancyUnit.unit_label == row.unit_code,
            TenancyUnit.deleted_at.is_(None),
        )
    ).first()
    if unit is None:
        result.blockers.append(
            f"No tenancy_unit with label={row.unit_code!r} under property {row.property_code!r}"
        )
        return result
    result.tenancy_unit_id = unit.id

    # Tenant by entity + exact legal_name.
    tenant = session.scalars(
        select(Tenant).where(
            Tenant.entity_id == entity_id,
            Tenant.legal_name == row.legal_name,
            Tenant.deleted_at.is_(None),
        )
    ).first()
    if tenant is None:
        result.blockers.append(f"No tenant with legal_name={row.legal_name!r}")
        return result
    result.tenant = tenant

    # Lease that links this tenant to this tenancy_unit.
    lease = session.scalars(
        select(Lease).where(
            Lease.tenancy_unit_id == unit.id,
            Lease.tenant_id == tenant.id,
            Lease.deleted_at.is_(None),
        )
    ).first()
    if lease is None:
        result.blockers.append(
            "No lease found linking this tenant to this unit"
        )
        return result
    result.lease = lease

    # Existing onboarding (if any) — pick the latest non-cancelled.
    onboarding = session.scalars(
        select(TenantOnboarding)
        .where(
            TenantOnboarding.tenant_id == tenant.id,
            TenantOnboarding.lease_id == lease.id,
            TenantOnboarding.deleted_at.is_(None),
            TenantOnboarding.status != TenantOnboardingStatus.cancelled,
        )
        .order_by(TenantOnboarding.created_at.desc())
    ).first()
    result.onboarding = onboarding
    return result


def _synthetic_token(tenant_id: UUID, lease_id: UUID) -> str:
    """Deterministic synthetic token for the back-fill onboarding record.

    The real onboarding flow generates random tokens for portal links.
    This record never had a portal link — it's a record of "this tenant
    was already enrolled before Relby". A deterministic hash means
    re-running the script doesn't create duplicate records.
    """
    seed = f"{ENROLLMENT_SOURCE}:{tenant_id}:{lease_id}".encode()
    digest = hashlib.sha256(seed).hexdigest()[:48]
    return f"already-enrolled-{digest}"


def _tenant_field_updates(tenant: Tenant, row: TenancyRow) -> dict[str, Any]:
    """Compute which tenant fields are blank and should be filled from row."""
    updates: dict[str, Any] = {}
    # Only fill blanks; never overwrite existing data.
    if not tenant.trading_name and row.trading_name:
        updates["trading_name"] = row.trading_name
    if not tenant.contact_email and _is_email(row.primary_contact):
        updates["contact_email"] = row.primary_contact
    if not tenant.notes and row.notes:
        updates["notes"] = row.notes
    return updates


def _plan_row(match: MatchResult) -> dict[str, Any]:
    """Build a per-row plan describing what would happen on apply."""
    plan: dict[str, Any] = {
        "row": match.row.sheet_row,
        "property": match.row.property_code,
        "unit": match.row.unit_code,
        "tenant": match.row.legal_name,
    }
    if match.blocked():
        plan["action"] = "blocked"
        plan["blockers"] = match.blockers
        return plan

    assert match.tenant is not None and match.lease is not None
    already_marked = (
        (match.tenant.tenant_metadata or {}).get("enrollment_source") == ENROLLMENT_SOURCE
        and match.onboarding is not None
        and match.onboarding.status == TenantOnboardingStatus.applied
        and (match.onboarding.review_data or {}).get("enrollment_source") == ENROLLMENT_SOURCE
    )
    if already_marked:
        plan["action"] = "skip"
        plan["reason"] = "already marked enrolled by a prior run"
        return plan

    tenant_updates = _tenant_field_updates(match.tenant, match.row)
    if match.onboarding is None:
        onboarding_action = "create (status=applied)"
    elif match.onboarding.status == TenantOnboardingStatus.applied:
        onboarding_action = "tag with enrollment_source (status already applied)"
    else:
        onboarding_action = f"transition {match.onboarding.status} -> applied"

    plan["action"] = "apply"
    plan["tenant_updates"] = tenant_updates or "(no blank fields to fill)"
    plan["onboarding_action"] = onboarding_action
    return plan


def _apply_row(
    session: Session,
    match: MatchResult,
    operator_user: AppUser,
    now: datetime,
) -> None:
    assert match.tenant is not None and match.lease is not None

    # Tenant updates (fill blanks only).
    updates = _tenant_field_updates(match.tenant, match.row)
    for field, value in updates.items():
        setattr(match.tenant, field, value)
    metadata = dict(match.tenant.tenant_metadata or {})
    metadata["enrollment_source"] = ENROLLMENT_SOURCE
    metadata["enrollment_applied_at"] = now.isoformat()
    metadata.setdefault(
        "enrollment_source_workbook_sheet",
        "Tenancies",
    )
    metadata["enrollment_source_row"] = match.row.sheet_row
    match.tenant.tenant_metadata = metadata

    # Onboarding upsert.
    review_payload = {
        "enrollment_source": ENROLLMENT_SOURCE,
        "enrollment_source_sheet": "Tenancies",
        "enrollment_source_row": match.row.sheet_row,
        "workbook_tenant_legal_name": match.row.legal_name,
        "workbook_trading_name": match.row.trading_name,
        "workbook_primary_contact": match.row.primary_contact,
        "workbook_status": match.row.status,
        "applied_by_email": operator_user.email,
    }

    if match.onboarding is None:
        new_onboarding = TenantOnboarding(
            entity_id=match.tenant.entity_id,
            lease_id=match.lease.id,
            tenant_id=match.tenant.id,
            token=_synthetic_token(match.tenant.id, match.lease.id),
            status=TenantOnboardingStatus.applied,
            submitted_data={},
            review_data=review_payload,
            delivery_data={
                "channel": "synthetic_already_enrolled",
                "message": "Marked enrolled from SKJ workbook; no portal link was issued.",
            },
            submitted_at=now,
            reviewed_at=now,
            reviewed_by_user_id=operator_user.id,
            applied_at=now,
            applied_by_user_id=operator_user.id,
        )
        session.add(new_onboarding)
    else:
        match.onboarding.status = TenantOnboardingStatus.applied
        match.onboarding.submitted_at = match.onboarding.submitted_at or now
        match.onboarding.reviewed_at = now
        match.onboarding.reviewed_by_user_id = (
            match.onboarding.reviewed_by_user_id or operator_user.id
        )
        match.onboarding.applied_at = now
        match.onboarding.applied_by_user_id = operator_user.id
        existing_review = dict(match.onboarding.review_data or {})
        existing_review.update(review_payload)
        match.onboarding.review_data = existing_review


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Mark active tenants from the SKJ portfolio workbook as already "
            "enrolled. No provider sends. Default dry-run; pass --apply to "
            "mutate."
        )
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        required=True,
        help="Path to the SKJ portfolio source-of-truth .xlsx workbook.",
    )
    parser.add_argument(
        "--entity-id",
        type=UUID,
        required=True,
        help="Entity UUID to scope the enrollment to.",
    )
    parser.add_argument(
        "--operator-email",
        required=True,
        help="Email of the operator applying the enrollment (must exist in app_user).",
    )
    parser.add_argument(
        "--database-url",
        default=None,
        help=(
            "Override the database URL. Defaults to the Settings() default "
            "(typically read from DATABASE_URL env var or .env)."
        ),
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually write to the database. Without this flag, runs as a dry-run.",
    )
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"ERROR: workbook not found at {args.workbook}", file=sys.stderr)
        return 2

    rows = _read_tenancies(args.workbook)
    if not rows:
        print("No enrollable rows found in workbook (after VACANT/inactive filter).")
        return 0

    database_url = (
        Settings(database_url=args.database_url).database_url
        if args.database_url
        else Settings().database_url
    )
    engine = create_engine(database_url, future=True, pool_pre_ping=True)
    SessionFactory = sessionmaker(bind=engine, expire_on_commit=False, future=True)

    with SessionFactory() as session:
        entity = session.get(Entity, args.entity_id)
        if entity is None:
            print(f"ERROR: entity {args.entity_id} not found", file=sys.stderr)
            return 2
        operator = session.scalars(
            select(AppUser).where(AppUser.email == args.operator_email)
        ).first()
        if operator is None:
            print(
                f"ERROR: operator email {args.operator_email!r} not found in app_user",
                file=sys.stderr,
            )
            return 2

        matches = [_match_row(session, args.entity_id, row) for row in rows]
        plans = [_plan_row(m) for m in matches]

        applies = [p for p in plans if p.get("action") == "apply"]
        skips = [p for p in plans if p.get("action") == "skip"]
        blocked = [p for p in plans if p.get("action") == "blocked"]

        print(f"Entity:         {entity.name} ({entity.id})")
        print(f"Operator:       {operator.email}")
        print(f"Workbook:       {args.workbook.name}")
        print(f"Source key:     {ENROLLMENT_SOURCE}")
        print(f"Rows considered (after VACANT/inactive filter): {len(rows)}")
        print(f"  - will apply:  {len(applies)}")
        print(f"  - will skip:   {len(skips)} (already marked by prior run)")
        print(f"  - blocked:     {len(blocked)} (could not match registers)")
        print()

        for plan in plans:
            tag = {
                "apply": "APPLY  ",
                "skip": "SKIP   ",
                "blocked": "BLOCKED",
            }.get(plan["action"], "?      ")
            print(
                f"  {tag}  row {plan['row']:>3}  "
                f"{plan['property']}/{plan['unit']}  {plan['tenant']}"
            )
            if plan["action"] == "apply":
                print(f"           tenant updates:   {plan['tenant_updates']}")
                print(f"           onboarding:       {plan['onboarding_action']}")
            elif plan["action"] == "blocked":
                for b in plan["blockers"]:
                    print(f"           - {b}")
            elif plan["action"] == "skip":
                print(f"           reason: {plan['reason']}")

        if not args.apply:
            print()
            print("Dry-run complete. Re-run with --apply to mutate.")
            return 0

        if not applies:
            print()
            print("Nothing to apply. Exiting cleanly.")
            return 0

        now = datetime.now(UTC)
        for match in matches:
            if match.blocked():
                continue
            plan = _plan_row(match)
            if plan["action"] != "apply":
                continue
            _apply_row(session, match, operator, now)

        session.commit()
        print()
        print(f"Applied enrollment for {len(applies)} tenant(s) at {now.isoformat()}.")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
