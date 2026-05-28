"""Review-first spreadsheet import routes."""

from io import BytesIO
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, Response, UploadFile, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session
from stewart.core.audit import audit_log
from stewart.core.db import utcnow
from stewart.core.models import RegisterImportPlan, UserRole
from stewart.domain.register_import import (
    RegisterImportError,
    apply_register_import_plan,
    build_register_import_dry_run,
)

from apps.api.deps import CurrentUser, assert_entity_role, get_current_user, get_session
from apps.api.schemas.register_import import (
    RegisterImportApplyRead,
    RegisterImportApplyRequest,
    RegisterImportDryRunRead,
)

router = APIRouter(prefix="/register-imports", tags=["register-imports"])

WRITE_ROLES = {UserRole.owner, UserRole.admin, UserRole.finance, UserRole.ops}
SUPPORTED_CONTENT_TYPES = {
    "application/octet-stream",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TEMPLATE_SHEETS: tuple[tuple[str, list[str]], ...] = (
    (
        "Entities",
        [
            "Entity (legal name)",
            "Type",
            "Role",
            "ABN / registration",
            "Billing email",
            "Xero contact name",
            "Properties",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Properties",
        [
            "Code",
            "Suburb",
            "Address",
            "Owning entity (legal)",
            "Role",
            "Property type",
            "Active tenancies",
            "Status",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Tenancies",
        [
            "Tenancy ID",
            "Property",
            "Unit code",
            "Tenant (legal name)",
            "Trading name",
            "Size m²",
            "Commencement",
            "Expiry",
            "Status",
            "Annual rent",
            "Outgoings",
            "Security",
            "Frequency",
            "Rent per m²",
            "Form",
            "Insurance",
            "Arrears",
            "Review type",
            "Next review",
            "Options",
            "Primary contact",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Charge Rules",
        [
            "Tenancy ID",
            "Charge type",
            "Amount",
            "Frequency",
            "GST treatment",
            "Start date",
            "End date",
            "Next due date",
            "Billing identity",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Bonds",
        [
            "Tenancy",
            "Property",
            "Tenant",
            "Security type",
            "Amount $AUD",
            "Paid date",
            "Security expiry",
            "Insurance status",
            "Insurance expiry",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Dates",
        [
            "Date",
            "Property",
            "Tenancy",
            "Event type",
            "Description",
            "Severity",
            "Owner",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Vendors",
        [
            "Vendor / Counterparty",
            "Category",
            "Scope",
            "Sites / properties",
            "Contact",
            "Email",
            "Phone",
            "Status",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Arrears",
        [
            "Tenancy ID",
            "Tenant",
            "Amount overdue",
            "Days overdue",
            "Last contact",
            "Promise to pay",
            "Owner",
            "Status",
            "Notes",
            "Source / confidence hint",
        ],
    ),
    (
        "Active Issues",
        [
            "Issue",
            "Property",
            "Tenancy",
            "Severity",
            "Owner",
            "Status",
            "Notes / next step",
            "Source / confidence hint",
        ],
    ),
    (
        "Actions",
        [
            "Deadline",
            "Action",
            "Detail",
            "Owner",
            "Property",
            "Tenancy",
            "Status",
            "Source / confidence hint",
        ],
    ),
)


def _validate_workbook_upload(filename: str, content_type: str | None) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Register imports currently support .xlsx workbooks.",
        )
    if content_type and content_type not in SUPPORTED_CONTENT_TYPES:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Register imports currently support .xlsx workbooks.",
        )


def _append_header_sheet(workbook: Workbook, title: str, headers: list[str]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6F5B")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(len(header) + 4, 14),
            32,
        )


def _migration_template_content() -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    assert instructions is not None
    instructions.title = "Instructions"
    instructions.append(["Leasium Smart Intake migration template"])
    instructions.append(["1. Download this template and keep the sheet names unchanged."])
    instructions.append(
        ["2. Complete it using your preferred AI or spreadsheet workflow."]
    )
    instructions.append(
        [
            "3. Upload the completed workbook back into Smart Intake for dry-run, "
            "row-level review, and explicit Apply."
        ]
    )
    instructions.append(
        [
            "4. Nothing changes in the register during template download or dry-run; "
            "only approved actions are applied."
        ]
    )
    instructions.append(
        [
            "5. Use Source / confidence hint for row provenance, extraction confidence, "
            "or notes from the migration source."
        ]
    )
    instructions.append([])
    instructions.append(
        [
            "Current Smart Intake applies Properties, Tenancies, Bonds, and Dates. "
            "Entities, Vendors, Charge Rules, Arrears, Active Issues, and Actions are "
            "kept visible for review and workflow staging."
        ]
    )
    instructions["A1"].font = Font(bold=True, size=14, color="1F6F5B")
    instructions.column_dimensions["A"].width = 118
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for title, headers in TEMPLATE_SHEETS:
        _append_header_sheet(workbook, title, headers)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@router.get("/template")
def download_register_import_template(
    _: Annotated[CurrentUser, Depends(get_current_user)],
) -> Response:
    """Return the Smart Intake migration template without creating an import plan."""

    return Response(
        content=_migration_template_content(),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": 'attachment; filename="leasium-migration-template.xlsx"'
        },
    )


@router.post("/dry-run", response_model=RegisterImportDryRunRead)
async def dry_run_register_import(
    user: Annotated[CurrentUser, Depends(get_current_user)],
    session: Annotated[Session, Depends(get_session)],
    entity_id: Annotated[UUID, Form()],
    file: Annotated[UploadFile, File()],
) -> RegisterImportDryRunRead:
    """Read a source-of-truth workbook and return a no-mutation import plan."""

    assert_entity_role(session, user, entity_id, WRITE_ROLES)
    _validate_workbook_upload(file.filename or "", file.content_type)
    content = await file.read()
    try:
        dry_run = build_register_import_dry_run(
            session=session,
            entity_id=entity_id,
            filename=file.filename or "register-import.xlsx",
            content=content,
        )
    except RegisterImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(exc),
        ) from exc
    dry_run_read = RegisterImportDryRunRead.model_validate(dry_run)
    plan = RegisterImportPlan(
        entity_id=entity_id,
        filename=dry_run_read.filename,
        plan_data=dry_run_read.model_dump(mode="json", exclude={"plan_id"}),
        created_by_user_id=user.id,
    )
    session.add(plan)
    session.commit()
    session.refresh(plan)
    return dry_run_read.model_copy(update={"plan_id": plan.id})


@router.post("/apply", response_model=RegisterImportApplyRead)
def apply_register_import(
    payload: RegisterImportApplyRequest,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    session: Annotated[Session, Depends(get_session)],
) -> RegisterImportApplyRead:
    """Apply only user-approved actions from a reviewed spreadsheet dry-run plan."""

    assert_entity_role(session, user, payload.entity_id, WRITE_ROLES)
    filename = payload.filename
    action_items = [item.model_dump(mode="json") for item in payload.action_items]
    plan: RegisterImportPlan | None = None
    if payload.plan_id is not None:
        plan = session.scalar(
            select(RegisterImportPlan).where(
                RegisterImportPlan.id == payload.plan_id,
                RegisterImportPlan.entity_id == payload.entity_id,
                RegisterImportPlan.deleted_at.is_(None),
            )
        )
        if plan is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Register import plan not found.",
            )
        if plan.applied_at is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Register import plan has already been applied.",
            )
        plan_data = dict(plan.plan_data or {})
        filename = str(plan_data.get("filename") or plan.filename)
        raw_action_items = plan_data.get("action_items")
        if not isinstance(raw_action_items, list):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Register import plan has no stored action items.",
            )
        action_items = raw_action_items

    try:
        result = apply_register_import_plan(
            session=session,
            entity_id=payload.entity_id,
            filename=filename,
            action_items=action_items,
            approved_action_ids=payload.approved_action_ids,
            ignored_action_ids=payload.ignored_action_ids,
        )
    except RegisterImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(exc),
        ) from exc

    for item in result["results"]:
        if item["status"] != "applied" or item.get("target_id") is None:
            continue
        audit_log(
            session,
            actor=user.actor,
            user_id=user.id,
            entity_id=payload.entity_id,
            action=item["operation"],
            target_table=item.get("target_table"),
            target_id=item["target_id"],
            tool_name="register_import_apply",
            tool_input={
                "filename": filename,
                "plan_id": str(payload.plan_id) if payload.plan_id else None,
                "action_id": item["action_id"],
            },
            tool_output_summary=item["message"],
        )
    result_read = RegisterImportApplyRead.model_validate(result)
    if plan is not None:
        plan.applied_at = result_read.applied_at
        plan.applied_by_user_id = user.id
        plan_data = dict(plan.plan_data or {})
        plan_data["apply_result"] = result_read.model_dump(mode="json")
        plan_data["approved_action_ids"] = list(payload.approved_action_ids)
        plan_data["ignored_action_ids"] = list(payload.ignored_action_ids)
        plan_data["notes"] = payload.notes
        plan.plan_data = plan_data
        plan.updated_at = utcnow()
    audit_log(
        session,
        actor=user.actor,
        user_id=user.id,
        entity_id=payload.entity_id,
        action="apply",
        tool_name="register_import_apply",
        tool_input={
            "filename": filename,
            "plan_id": str(payload.plan_id) if payload.plan_id else None,
            "approved_action_ids": payload.approved_action_ids,
            "ignored_action_ids": payload.ignored_action_ids,
        },
        tool_output_summary=(
            f"Applied {result['applied']} register import actions; "
            f"{result['blocked']} blocked and {result['skipped']} skipped."
        ),
    )
    session.commit()
    return result_read
