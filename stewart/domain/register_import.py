"""Review-first spreadsheet import planning for portfolio registers."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from stewart.core.db import utcnow
from stewart.core.models import (
    GstTreatment,
    Lease,
    LeaseStatus,
    Obligation,
    ObligationCategory,
    ObligationStatus,
    Property,
    PropertyType,
    RentChargeRule,
    RentChargeType,
    RentFrequency,
    TenancyUnit,
    Tenant,
    UserRole,
)


class RegisterImportError(ValueError):
    """Raised when a workbook cannot be read as a register import source."""


@dataclass(frozen=True)
class SheetRows:
    name: str
    header_row: int | None
    columns: list[str]
    rows: list[dict[str, Any]]


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"", "-", "—", "N/A", "n/a", "NA"}:
        return ""
    return text


def _key(value: Any) -> str:
    return " ".join(_text(value).lower().split())


def _money_cents(value: Any) -> int | None:
    if value in (None, "", "-", "—"):
        return None
    if isinstance(value, int | float):
        return int(round(float(value) * 100))
    cleaned = "".join(char for char in str(value) if char.isdigit() or char in ".-")
    if not cleaned:
        return None
    try:
        return int(round(float(cleaned) * 100))
    except ValueError:
        return None


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _float_value(value: Any) -> float | None:
    if value in (None, "", "-", "—"):
        return None
    if isinstance(value, int | float):
        return float(value)
    cleaned = "".join(char for char in str(value) if char.isdigit() or char in ".-")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _int_value(value: Any) -> int | None:
    number = _float_value(value)
    if number is None:
        return None
    return int(round(number))


def _rent_frequency(value: Any) -> RentFrequency | None:
    text = _key(value)
    if "week" in text:
        return RentFrequency.weekly
    if "month" in text:
        return RentFrequency.monthly
    if "quarter" in text:
        return RentFrequency.quarterly
    if "annual" in text or "year" in text or "annum" in text or text == "pa":
        return RentFrequency.annual
    return None


def _periodic_rent_cents(annual_rent_cents: int, frequency: RentFrequency) -> int:
    divisor = 1
    if frequency == RentFrequency.weekly:
        divisor = 52
    elif frequency == RentFrequency.monthly:
        divisor = 12
    elif frequency == RentFrequency.quarterly:
        divisor = 4
    return int(round(annual_rent_cents / divisor))


def _jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if hasattr(value, "value"):
        return value.value
    if isinstance(value, dict):
        return {key: _jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_jsonable(item) for item in value]
    return value


def _blank_to_none(value: Any) -> Any:
    return None if _text(value) == "" else value


def _sheet_rows(workbook: Workbook, sheet_name: str) -> SheetRows:
    sheet = workbook[sheet_name]
    header_row: int | None = None
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        values = [
            _text(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)
        ]
        if sum(1 for value in values if value) >= 2:
            header_row = row_number
            break

    if header_row is None:
        return SheetRows(name=sheet_name, header_row=None, columns=[], rows=[])

    column_positions = [
        (column, _text(sheet.cell(header_row, column).value))
        for column in range(1, sheet.max_column + 1)
        if _text(sheet.cell(header_row, column).value)
    ]
    rows: list[dict[str, Any]] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        row = {
            header: _blank_to_none(sheet.cell(row_number, column).value)
            for column, header in column_positions
        }
        if any(_text(value) for value in row.values()):
            row["_row"] = row_number
            rows.append(row)
    return SheetRows(
        name=sheet_name,
        header_row=header_row,
        columns=[header for _, header in column_positions],
        rows=rows,
    )


def load_register_workbook(content: bytes) -> dict[str, SheetRows]:
    """Load an XLSX workbook into normalized sheet rows."""

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - library message varies by file corruption
        raise RegisterImportError("Could not read workbook. Upload a valid .xlsx file.") from exc
    return {sheet_name: _sheet_rows(workbook, sheet_name) for sheet_name in workbook.sheetnames}


def _property_type(value: Any) -> tuple[PropertyType, bool]:
    text = _key(value)
    if not text or text == "tbc":
        return PropertyType.other, True
    if "childcare" in text:
        return PropertyType.childcare, False
    if "hospitality" in text or "bottle shop" in text:
        return PropertyType.hospitality, False
    if "industrial" in text or "logistics" in text or "brewery" in text:
        return PropertyType.commercial_industrial, False
    if "office" in text or "hq" in text:
        return PropertyType.commercial_office, False
    if "retail" in text:
        return PropertyType.commercial_retail, False
    if "mixed" in text:
        return PropertyType.mixed_use, False
    if "land" in text:
        return PropertyType.vacant_land, False
    if "residential" in text:
        return PropertyType.other, True
    if "commercial" in text:
        return PropertyType.commercial_retail, True
    return PropertyType.other, True


def _lease_status_bucket(value: Any) -> str:
    status = _key(value)
    if not status:
        return "review"
    if "vacant" in status:
        return "vacant"
    if status.startswith("exited"):
        return "archive"
    if "skj-as-tenant" in status:
        return "headlease"
    if status.startswith("pending"):
        return "pending"
    if status.startswith("active"):
        return "active"
    return "review"


def _has_arrears_signal(status: Any, arrears: Any) -> bool:
    if "arrears" in _key(status):
        return True
    text = _key(arrears)
    return text not in {"", "not specified", "none", "nil", "no", "current"}


def _obligation_category(event_type: Any) -> str:
    text = _key(event_type)
    if "review" in text or "cpi" in text:
        return "rent_review"
    if "expiry" in text or "renewal" in text:
        return "lease_expiry"
    if "insurance" in text:
        return "insurance"
    if "guarantee" in text or "bond" in text:
        return "bank_guarantee"
    if "compliance" in text or "regulatory" in text:
        return "compliance"
    if "maintenance" in text or "leak" in text or "vendor" in text:
        return "maintenance"
    return "other"


def _finding(
    severity: str,
    message: str,
    *,
    sheet: str | None = None,
    row: int | None = None,
    field: str | None = None,
    source_value: Any = None,
) -> dict[str, Any]:
    return {
        "severity": severity,
        "message": message,
        "sheet": sheet,
        "row": row,
        "field": field,
        "source_value": source_value,
    }


def _feature(
    key: str,
    label: str,
    reason: str,
    source_sheet: str,
    source_count: int,
    priority: str = "next",
) -> dict[str, Any]:
    return {
        "key": key,
        "label": label,
        "reason": reason,
        "source_sheet": source_sheet,
        "source_count": source_count,
        "priority": priority,
    }


def _action(target: str, **counts: int) -> dict[str, int | str]:
    return {
        "target": target,
        "create": counts.get("create", 0),
        "match": counts.get("match", 0),
        "update": counts.get("update", 0),
        "skip": counts.get("skip", 0),
        "review": counts.get("review", 0),
    }


def _source_context(
    filename: str,
    sheet: str,
    row: int | None,
    *,
    confidence: float | None = 0.86,
) -> dict[str, Any]:
    source_hint = sheet
    if row is not None:
        source_hint = f"{sheet} row {row}"
    return {
        "filename": filename,
        "sheet": sheet,
        "row": row,
        "source_hint": source_hint,
        "confidence": confidence,
    }


def _change(
    filename: str,
    sheet: str,
    row: int | None,
    field: str,
    label: str,
    before: Any,
    after: Any,
) -> dict[str, Any]:
    return {
        "field": field,
        "label": label,
        "before": _jsonable(before),
        "after": _jsonable(after),
        "source": _source_context(filename, sheet, row),
    }


def _action_item(
    *,
    action_id: str,
    target: str,
    operation: str,
    label: str,
    summary: str,
    source: dict[str, Any],
    payload: dict[str, Any],
    changes: list[dict[str, Any]] | None = None,
    blockers: list[str] | None = None,
    warnings: list[str] | None = None,
    default_decision: str | None = None,
) -> dict[str, Any]:
    item_blockers = blockers or []
    item_warnings = warnings or []
    decision = default_decision
    if decision is None:
        if item_blockers or operation in {"review", "skip", "match"}:
            decision = "review" if item_blockers else "ignore"
        else:
            decision = "approve"
    return {
        "id": action_id,
        "target": target,
        "operation": operation,
        "label": label,
        "summary": summary,
        "source": source,
        "changes": changes or [],
        "payload": _jsonable(payload),
        "blockers": item_blockers,
        "warnings": item_warnings,
        "default_decision": decision,
    }


def _existing_indexes(session: Session, entity_id: UUID) -> dict[str, Any]:
    properties = list(
        session.scalars(
            select(Property).where(Property.entity_id == entity_id, Property.deleted_at.is_(None))
        )
    )
    property_by_code = {
        _key((prop.property_metadata or {}).get("portfolio_code")): prop
        for prop in properties
        if _key((prop.property_metadata or {}).get("portfolio_code"))
    }
    property_by_address = {
        _key(prop.street_address): prop for prop in properties if _key(prop.street_address)
    }
    property_codes = {
        _key((prop.property_metadata or {}).get("portfolio_code"))
        for prop in properties
        if _key((prop.property_metadata or {}).get("portfolio_code"))
    }
    property_addresses = {
        _key(prop.street_address) for prop in properties if _key(prop.street_address)
    }
    tenants = list(
        session.scalars(
            select(Tenant).where(Tenant.entity_id == entity_id, Tenant.deleted_at.is_(None))
        )
    )
    tenant_by_name = {_key(tenant.legal_name): tenant for tenant in tenants}
    tenant_by_abn = {_key(tenant.abn): tenant for tenant in tenants if _key(tenant.abn)}
    tenant_names = {_key(tenant.legal_name) for tenant in tenants}
    tenant_abns = {_key(tenant.abn) for tenant in tenants if _key(tenant.abn)}

    units = list(
        session.scalars(
            select(TenancyUnit)
            .join(Property)
            .where(
                Property.entity_id == entity_id,
                Property.deleted_at.is_(None),
                TenancyUnit.deleted_at.is_(None),
            )
        )
    )
    unit_labels = {_key(unit.unit_label) for unit in units}
    unit_by_property_and_label = {
        (str(unit.property_id), _key(unit.unit_label)): unit
        for unit in units
        if _key(unit.unit_label)
    }
    leases = list(
        session.scalars(
            select(Lease)
            .join(TenancyUnit)
            .join(Property)
            .where(
                Property.entity_id == entity_id,
                Property.deleted_at.is_(None),
                TenancyUnit.deleted_at.is_(None),
                Lease.deleted_at.is_(None),
            )
        )
    )
    lease_refs = {
        _key((lease.lease_metadata or {}).get("portfolio_tenancy_id"))
        for lease in leases
        if _key((lease.lease_metadata or {}).get("portfolio_tenancy_id"))
    }
    lease_by_ref = {
        _key((lease.lease_metadata or {}).get("portfolio_tenancy_id")): lease
        for lease in leases
        if _key((lease.lease_metadata or {}).get("portfolio_tenancy_id"))
    }
    return {
        "property_codes": property_codes,
        "property_addresses": property_addresses,
        "property_by_code": property_by_code,
        "property_by_address": property_by_address,
        "tenant_names": tenant_names,
        "tenant_abns": tenant_abns,
        "tenant_by_name": tenant_by_name,
        "tenant_by_abn": tenant_by_abn,
        "unit_labels": unit_labels,
        "unit_by_property_and_label": unit_by_property_and_label,
        "lease_refs": lease_refs,
        "lease_by_ref": lease_by_ref,
    }


def _safe_action_key(*parts: Any) -> str:
    cleaned = [
        "-".join(char for char in _key(part).replace("/", " ").split() if char)
        for part in parts
        if _text(part)
    ]
    return ":".join(cleaned) or "row"


def _property_payload_from_row(row: dict[str, Any]) -> dict[str, Any]:
    code = _text(row.get("Code"))
    suburb = _text(row.get("Suburb"))
    address = _text(row.get("Address"))
    property_type, type_needs_review = _property_type(row.get("Property type"))
    return {
        "portfolio_code": code,
        "name": code or suburb or address or "Property to confirm",
        "street_address": address,
        "suburb": suburb or None,
        "country_code": "AU",
        "property_type": property_type.value,
        "ownership_structure": _text(row.get("Role")) or None,
        "owner_legal_name": _text(row.get("Owning entity (legal)")) or None,
        "invoice_reference": code or None,
        "active_tenancies": _int_value(row.get("Active tenancies")),
        "source_notes": _text(row.get("Notes")) or None,
        "source_property_type": _text(row.get("Property type")) or None,
        "property_type_needs_review": type_needs_review,
    }


def _property_create_changes(
    filename: str,
    row: dict[str, Any],
    payload: dict[str, Any],
) -> list[dict[str, Any]]:
    row_number = row.get("_row")
    fields = [
        ("name", "Name"),
        ("street_address", "Street address"),
        ("suburb", "Suburb"),
        ("property_type", "Property type"),
        ("ownership_structure", "Role"),
        ("owner_legal_name", "Owning entity"),
        ("invoice_reference", "Invoice reference"),
    ]
    return [
        _change(filename, "Properties", row_number, field, label, None, payload.get(field))
        for field, label in fields
        if payload.get(field) not in (None, "")
    ]


def _property_update_changes(
    filename: str,
    row: dict[str, Any],
    prop: Property,
    payload: dict[str, Any],
) -> list[dict[str, Any]]:
    row_number = row.get("_row")
    candidates = [
        ("suburb", "Suburb"),
        ("ownership_structure", "Role"),
        ("owner_legal_name", "Owning entity"),
        ("invoice_reference", "Invoice reference"),
    ]
    changes: list[dict[str, Any]] = []
    for field, label in candidates:
        after = payload.get(field)
        before = getattr(prop, field)
        if after not in (None, "") and before in (None, ""):
            changes.append(_change(filename, "Properties", row_number, field, label, before, after))
    if prop.property_type == PropertyType.other and payload.get("property_type"):
        after_type = payload["property_type"]
        if after_type != prop.property_type.value:
            changes.append(
                _change(
                    filename,
                    "Properties",
                    row_number,
                    "property_type",
                    "Property type",
                    prop.property_type,
                    after_type,
                )
            )
    return changes


def _matching_property_from_indexes(
    indexes: dict[str, Any],
    *,
    code: str,
    address: str,
) -> Property | None:
    if _key(code):
        prop = indexes["property_by_code"].get(_key(code))
        if prop is not None:
            return prop
    if _key(address):
        return indexes["property_by_address"].get(_key(address))
    return None


def _property_action_item(
    filename: str,
    row: dict[str, Any],
    indexes: dict[str, Any],
) -> dict[str, Any]:
    row_number = row.get("_row")
    source = _source_context(filename, "Properties", row_number)
    payload = _property_payload_from_row(row)
    code = payload["portfolio_code"]
    address = payload["street_address"]
    warnings: list[str] = []
    blockers: list[str] = []
    if not code:
        blockers.append("Property row is missing a portfolio code.")
    if not address:
        blockers.append("Property row is missing a street address.")
    if payload.pop("property_type_needs_review"):
        warnings.append("Property type should be checked before apply.")
    if "inactive" in _key(row.get("Status")) or "sold" in _key(row.get("Property type")):
        return _action_item(
            action_id=f"property:{row_number}:skip",
            target="properties",
            operation="skip",
            label=code or address or f"Property row {row_number}",
            summary="Inactive or sold property left out of v1 apply.",
            source=source,
            payload=payload,
            warnings=warnings,
            default_decision="ignore",
        )

    existing = _matching_property_from_indexes(indexes, code=code, address=address)
    if existing is not None:
        changes = _property_update_changes(filename, row, existing, payload)
        if changes:
            return _action_item(
                action_id=f"property:{row_number}:{_safe_action_key(code, address)}:update",
                target="properties",
                operation="update",
                label=existing.name,
                summary="Fill blank property fields from the workbook.",
                source=source,
                payload={
                    **payload,
                    "match": {"portfolio_code": code, "street_address": address},
                },
                changes=changes,
                warnings=warnings,
            )
        return _action_item(
            action_id=f"property:{row_number}:{_safe_action_key(code, address)}:match",
            target="properties",
            operation="match",
            label=existing.name,
            summary="Existing property matched; no field changes proposed.",
            source=source,
            payload={
                **payload,
                "match": {"portfolio_code": code, "street_address": address},
            },
            warnings=warnings,
            default_decision="ignore",
        )

    return _action_item(
        action_id=f"property:{row_number}:{_safe_action_key(code, address)}:create",
        target="properties",
        operation="create",
        label=code or address or f"Property row {row_number}",
        summary="Create property register record.",
        source=source,
        payload=payload,
        changes=_property_create_changes(filename, row, payload),
        blockers=blockers,
        warnings=warnings,
    )


def _tenancy_payload_from_row(row: dict[str, Any]) -> dict[str, Any]:
    tenancy_id = _text(row.get("Tenancy ID"))
    unit_label = _text(row.get("Unit code")) or tenancy_id
    frequency = _rent_frequency(row.get("Frequency"))
    annual_rent_cents = _money_cents(row.get("Annual rent"))
    return {
        "portfolio_tenancy_id": tenancy_id,
        "property_code": _text(row.get("Property")),
        "unit_label": unit_label,
        "tenant_legal_name": _text(row.get("Tenant (legal name)")),
        "tenant_trading_name": _text(row.get("Trading name")) or None,
        "unit_sqm": _float_value(row.get("Size m²")),
        "contact_name": _text(row.get("Primary contact")) or None,
        "commencement_date": _jsonable(_date_value(row.get("Commencement"))),
        "expiry_date": _jsonable(_date_value(row.get("Expiry"))),
        "status_bucket": _lease_status_bucket(row.get("Status")),
        "lease_status": (
            LeaseStatus.active.value
            if _lease_status_bucket(row.get("Status")) == "active"
            else LeaseStatus.pending.value
        ),
        "annual_rent_cents": annual_rent_cents,
        "rent_frequency": frequency.value if frequency is not None else None,
        "outgoings_amount_cents": _money_cents(row.get("Outgoings")),
        "next_review_date": _jsonable(_date_value(row.get("Next review"))),
        "review_type": _text(row.get("Review type")) or None,
        "option_summary": _text(row.get("Options")) or None,
        "security_summary": _text(row.get("Security")) or None,
        "insurance_status": _text(row.get("Insurance")) or None,
        "arrears": _text(row.get("Arrears")) or None,
        "form": _text(row.get("Form")) or None,
        "notes": _text(row.get("Notes")) or None,
    }


def _tenancy_action_item(
    filename: str,
    row: dict[str, Any],
    property_codes: set[str],
    tenancy_id_counts: Counter[str],
    indexes: dict[str, Any],
) -> dict[str, Any]:
    row_number = row.get("_row")
    source = _source_context(filename, "Tenancies", row_number)
    payload = _tenancy_payload_from_row(row)
    tenancy_id = payload["portfolio_tenancy_id"]
    bucket = payload["status_bucket"]
    blockers: list[str] = []
    warnings: list[str] = []
    if tenancy_id and tenancy_id_counts[tenancy_id] > 1:
        blockers.append("Tenancy ID appears more than once.")
    if not tenancy_id:
        blockers.append("Tenancy row is missing a Tenancy ID.")
    if payload["property_code"] not in property_codes:
        blockers.append("Tenancy references a property code not present in the Properties sheet.")
    if not _text(row.get("Unit code")):
        warnings.append("No unit code supplied; Tenancy ID will be used as the unit label.")
    if bucket == "vacant":
        return _action_item(
            action_id=f"tenancy:{row_number}:{_safe_action_key(tenancy_id)}:skip",
            target="tenancies",
            operation="skip",
            label=tenancy_id or f"Tenancy row {row_number}",
            summary="Vacant row left out of v1 apply.",
            source=source,
            payload=payload,
            warnings=warnings,
            default_decision="ignore",
        )
    if bucket in {"archive", "headlease", "review"}:
        blockers.append("Tenancy status needs human modelling before v1 apply.")
    if not payload["tenant_legal_name"] or str(payload["tenant_legal_name"]).upper() == "VACANT":
        blockers.append("Importable tenancy needs a tenant legal name.")
    if not payload["commencement_date"] or not payload["expiry_date"]:
        blockers.append("Importable tenancy needs commencement and expiry dates.")
    if payload["annual_rent_cents"] is None:
        blockers.append("Importable tenancy needs annual rent.")
    if payload["rent_frequency"] is None:
        warnings.append("Rent frequency is missing; monthly will be used for charge rules.")
        payload["rent_frequency"] = RentFrequency.monthly.value
    if _key(tenancy_id) in indexes["lease_refs"]:
        blockers.append("Lease with this portfolio tenancy ID already exists.")

    changes = [
        _change(
            filename,
            "Tenancies",
            row_number,
            "unit_label",
            "Unit",
            None,
            payload["unit_label"],
        ),
        _change(
            filename,
            "Tenancies",
            row_number,
            "tenant_legal_name",
            "Tenant",
            None,
            payload["tenant_legal_name"],
        ),
        _change(
            filename,
            "Tenancies",
            row_number,
            "annual_rent_cents",
            "Annual rent",
            None,
            payload["annual_rent_cents"],
        ),
    ]
    return _action_item(
        action_id=f"tenancy:{row_number}:{_safe_action_key(tenancy_id)}:create",
        target="tenancies",
        operation="create" if not blockers else "review",
        label=tenancy_id or f"Tenancy row {row_number}",
        summary="Create unit, tenant link, lease, and rent rules.",
        source=source,
        payload=payload,
        changes=[change for change in changes if change["after"] not in (None, "")],
        blockers=blockers,
        warnings=warnings,
    )


def _obligation_action_item(
    *,
    filename: str,
    sheet: str,
    row: dict[str, Any],
    action_key: str,
    title: str,
    due_date: date | None,
    category: str,
    property_code: str | None = None,
    tenancy_ref: str | None = None,
    notes: str | None = None,
    owner_role: str | None = None,
    blockers: list[str] | None = None,
) -> dict[str, Any]:
    row_number = row.get("_row")
    payload = {
        "title": title,
        "due_date": _jsonable(due_date),
        "category": category,
        "property_code": property_code,
        "tenancy_ref": tenancy_ref,
        "notes": notes,
        "owner_role": owner_role,
    }
    item_blockers = list(blockers or [])
    if due_date is None:
        item_blockers.append("Obligation needs a concrete due date before apply.")
    changes = [
        _change(filename, sheet, row_number, "title", "Title", None, title),
        _change(filename, sheet, row_number, "due_date", "Due date", None, due_date),
    ]
    return _action_item(
        action_id=f"obligation:{row_number}:{_safe_action_key(sheet, action_key)}:create",
        target="obligations",
        operation="create" if not item_blockers else "review",
        label=title,
        summary="Create obligation from workbook date context.",
        source=_source_context(filename, sheet, row_number),
        payload=payload,
        changes=[change for change in changes if change["after"] not in (None, "")],
        blockers=item_blockers,
    )


def build_register_import_dry_run(
    *,
    session: Session,
    entity_id: UUID,
    filename: str,
    content: bytes,
) -> dict[str, Any]:
    """Return a no-mutation import plan for a portfolio workbook."""

    sheets = load_register_workbook(content)
    indexes = _existing_indexes(session, entity_id)
    findings: list[dict[str, Any]] = []
    feature_candidates: list[dict[str, Any]] = []
    action_items: list[dict[str, Any]] = []

    properties = sheets.get("Properties", SheetRows("Properties", None, [], [])).rows
    tenancies = sheets.get("Tenancies", SheetRows("Tenancies", None, [], [])).rows
    bonds = sheets.get("Bonds", SheetRows("Bonds", None, [], [])).rows
    dates = sheets.get("Dates", SheetRows("Dates", None, [], [])).rows
    entities = sheets.get("Entities", SheetRows("Entities", None, [], [])).rows
    vendors = sheets.get("Vendors", SheetRows("Vendors", None, [], [])).rows
    charge_rules = sheets.get("Charge Rules", SheetRows("Charge Rules", None, [], [])).rows
    arrears_sheet = sheets.get("Arrears", SheetRows("Arrears", None, [], [])).rows
    issues = sheets.get("Active Issues", SheetRows("Active Issues", None, [], [])).rows
    actions_rows = sheets.get("Actions", SheetRows("Actions", None, [], [])).rows

    property_codes = {_text(row.get("Code")) for row in properties if _text(row.get("Code"))}
    tenancy_ids = [
        _text(row.get("Tenancy ID")) for row in tenancies if _text(row.get("Tenancy ID"))
    ]
    tenancy_id_counts = Counter(tenancy_ids)

    property_create = 0
    property_match = 0
    property_review = 0
    inactive_properties = 0
    for row in properties:
        action_items.append(_property_action_item(filename, row, indexes))
        code = _text(row.get("Code"))
        address = _text(row.get("Address"))
        status = _key(row.get("Status"))
        if not code:
            property_review += 1
            findings.append(
                _finding(
                    "blocker",
                    "Property row is missing a portfolio code.",
                    sheet="Properties",
                    row=row.get("_row"),
                    field="Code",
                )
            )
        if not address:
            property_review += 1
            findings.append(
                _finding(
                    "blocker",
                    "Property row is missing a street address.",
                    sheet="Properties",
                    row=row.get("_row"),
                    field="Address",
                )
            )
        _, needs_type_review = _property_type(row.get("Property type"))
        if needs_type_review:
            property_review += 1
            findings.append(
                _finding(
                    "warning",
                    "Property type needs human mapping before apply.",
                    sheet="Properties",
                    row=row.get("_row"),
                    field="Property type",
                    source_value=row.get("Property type"),
                )
            )
        if "inactive" in status or "sold" in _key(row.get("Property type")):
            inactive_properties += 1
            property_review += 1
            continue
        if (
            _key(code) in indexes["property_codes"]
            or _key(address) in indexes["property_addresses"]
        ):
            property_match += 1
        elif code and address:
            property_create += 1

    unit_create = 0
    unit_match = 0
    unit_review = 0
    tenant_names: set[str] = set()
    lease_create = 0
    lease_review = 0
    lease_skip = 0
    rent_rules = 0
    outgoings_rules = 0
    headlease_rows = 0
    vacant_rows = 0
    archive_rows = 0
    arrears_rows = 0

    for row in tenancies:
        action_items.append(
            _tenancy_action_item(filename, row, property_codes, tenancy_id_counts, indexes)
        )
        tenancy_id = _text(row.get("Tenancy ID"))
        prop_code = _text(row.get("Property"))
        unit_code = _text(row.get("Unit code"))
        tenant_name = _text(row.get("Tenant (legal name)"))
        bucket = _lease_status_bucket(row.get("Status"))

        if tenancy_id and tenancy_id_counts[tenancy_id] > 1:
            findings.append(
                _finding(
                    "blocker",
                    "Tenancy ID appears more than once.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Tenancy ID",
                    source_value=tenancy_id,
                )
            )
            lease_review += 1
            continue
        if not tenancy_id:
            findings.append(
                _finding(
                    "blocker",
                    "Tenancy row is missing a Tenancy ID.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Tenancy ID",
                )
            )
            lease_review += 1
            continue
        if prop_code not in property_codes:
            findings.append(
                _finding(
                    "blocker",
                    "Tenancy references a property code not present in the Properties sheet.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Property",
                    source_value=prop_code,
                )
            )
            lease_review += 1
            continue
        if not unit_code:
            findings.append(
                _finding(
                    "warning",
                    "Tenancy has no unit code; the importer can fall back "
                    "to Tenancy ID for the unit label.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Unit code",
                    source_value=tenancy_id,
                )
            )
            unit_review += 1
            unit_code = tenancy_id

        if _key(unit_code) in indexes["unit_labels"]:
            unit_match += 1
        else:
            unit_create += 1

        if bucket == "vacant":
            vacant_rows += 1
            lease_skip += 1
            continue
        if bucket == "archive":
            archive_rows += 1
            lease_review += 1
            continue
        if bucket == "headlease":
            headlease_rows += 1
            lease_review += 1
            continue
        if bucket == "review":
            lease_review += 1
            findings.append(
                _finding(
                    "warning",
                    "Tenancy status needs human mapping.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Status",
                    source_value=row.get("Status"),
                )
            )
            continue

        if not tenant_name or tenant_name.upper() == "VACANT":
            lease_review += 1
            findings.append(
                _finding(
                    "blocker",
                    "Importable lease row needs a tenant legal name.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Tenant (legal name)",
                )
            )
            continue
        if not _date_value(row.get("Commencement")) or not _date_value(row.get("Expiry")):
            lease_review += 1
            findings.append(
                _finding(
                    "blocker",
                    "Importable lease row needs commencement and expiry dates.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Commencement/Expiry",
                )
            )
            continue
        if _money_cents(row.get("Annual rent")) is None:
            lease_review += 1
            findings.append(
                _finding(
                    "blocker",
                    "Importable lease row needs annual rent.",
                    sheet="Tenancies",
                    row=row.get("_row"),
                    field="Annual rent",
                )
            )
            continue

        if _has_arrears_signal(row.get("Status"), row.get("Arrears")):
            arrears_rows += 1
        tenant_names.add(_key(tenant_name))
        if _key(tenancy_id) in indexes["lease_refs"]:
            lease_review += 1
        else:
            lease_create += 1
        rent_rules += 1
        if _money_cents(row.get("Outgoings")) is not None:
            outgoings_rules += 1

    tenant_match = sum(1 for name in tenant_names if name in indexes["tenant_names"])
    tenant_create = len(tenant_names) - tenant_match

    bond_security_obligations = 0
    bond_insurance_obligations = 0
    for row in bonds:
        tenancy_ref = _text(row.get("Tenancy"))
        prop_ref = _text(row.get("Property"))
        if tenancy_ref and tenancy_ref not in tenancy_id_counts:
            findings.append(
                _finding(
                    "blocker",
                    "Bond row references a tenancy not present in the Tenancies sheet.",
                    sheet="Bonds",
                    row=row.get("_row"),
                    field="Tenancy",
                    source_value=tenancy_ref,
                )
            )
        if prop_ref and prop_ref not in property_codes:
            findings.append(
                _finding(
                    "blocker",
                    "Bond row references a property not present in the Properties sheet.",
                    sheet="Bonds",
                    row=row.get("_row"),
                    field="Property",
                    source_value=prop_ref,
                )
            )
        if _date_value(row.get("Security expiry")) or _text(row.get("Security expiry")):
            bond_security_obligations += 1
            action_items.append(
                _obligation_action_item(
                    filename=filename,
                    sheet="Bonds",
                    row=row,
                    action_key=f"{tenancy_ref}-security",
                    title=f"Confirm security for {tenancy_ref or prop_ref or 'tenancy'}",
                    due_date=_date_value(row.get("Security expiry")),
                    category=ObligationCategory.bank_guarantee.value,
                    property_code=prop_ref or None,
                    tenancy_ref=tenancy_ref or None,
                    notes=_text(row.get("Notes")) or _text(row.get("Security type")) or None,
                    blockers=[],
                )
            )
        if _date_value(row.get("Insurance expiry")):
            bond_insurance_obligations += 1
            action_items.append(
                _obligation_action_item(
                    filename=filename,
                    sheet="Bonds",
                    row=row,
                    action_key=f"{tenancy_ref}-insurance",
                    title=f"Insurance expiry for {tenancy_ref or prop_ref or 'tenancy'}",
                    due_date=_date_value(row.get("Insurance expiry")),
                    category=ObligationCategory.insurance.value,
                    property_code=prop_ref or None,
                    tenancy_ref=tenancy_ref or None,
                    notes=_text(row.get("Insurance status")) or _text(row.get("Notes")) or None,
                    blockers=[],
                )
            )

    date_obligations = 0
    for row in dates:
        prop_ref = _text(row.get("Property"))
        tenancy_ref = _text(row.get("Tenancy"))
        if prop_ref and prop_ref not in property_codes:
            findings.append(
                _finding(
                    "blocker",
                    "Critical date references a property not present in the Properties sheet.",
                    sheet="Dates",
                    row=row.get("_row"),
                    field="Property",
                    source_value=prop_ref,
                )
            )
            continue
        if _date_value(row.get("Date")) is None:
            findings.append(
                _finding(
                    "blocker",
                    "Critical date row needs a date before it can become an obligation.",
                    sheet="Dates",
                    row=row.get("_row"),
                    field="Date",
                )
            )
            continue
        _obligation_category(row.get("Event type"))
        date_obligations += 1
        action_items.append(
            _obligation_action_item(
                filename=filename,
                sheet="Dates",
                row=row,
                action_key=f"{prop_ref}-{tenancy_ref}-{_text(row.get('Event type'))}",
                title=_text(row.get("Description"))
                or _text(row.get("Event type"))
                or f"Critical date for {prop_ref or tenancy_ref or 'portfolio'}",
                due_date=_date_value(row.get("Date")),
                category=_obligation_category(row.get("Event type")),
                property_code=prop_ref or None,
                tenancy_ref=tenancy_ref or None,
                notes=_text(row.get("Description")) or None,
                owner_role=_key(row.get("Owner")) or None,
                blockers=[],
            )
        )

    if entities:
        feature_candidates.append(
            _feature(
                "legal_entity_directory",
                "Legal entity and trust directory",
                "The workbook tracks owning/trading entities separately from "
                "properties; this should become a richer entity profile import.",
                "Entities",
                len(entities),
                "now",
            )
        )
    if vendors:
        feature_candidates.append(
            _feature(
                "vendor_directory",
                "Vendor and contractor directory",
                "Vendor scope, property coverage, and status appear as "
                "source-of-truth data but Leasium does not yet have a vendor register.",
                "Vendors",
                len(vendors),
            )
        )
    if charge_rules:
        feature_candidates.append(
            _feature(
                "charge_rule_review",
                "Charge rule migration review",
                "The workbook has standalone charge-rule rows that should be "
                "reviewed against lease charge rules before import mapping is expanded.",
                "Charge Rules",
                len(charge_rules),
                "next",
            )
        )
    if issues or actions_rows:
        feature_candidates.append(
            _feature(
                "issue_action_queue",
                "Active issue and action queue",
                "The workbook has operational issues and deadlines that "
                "should feed tasks, owners, severity, and follow-up views.",
                "Active Issues / Actions",
                len(issues) + len(actions_rows),
                "now",
            )
        )
    if bonds:
        feature_candidates.append(
            _feature(
                "security_originals_register",
                "Security originals and insurance chase register",
                "Bank guarantees, cash bonds, original-document locations, "
                "insurance expiry, and chase states need structured tracking.",
                "Bonds",
                len(bonds),
                "now",
            )
        )
    if headlease_rows:
        feature_candidates.append(
            _feature(
                "headlease_role_model",
                "SKJ-as-tenant and headlease modelling",
                "Some rows describe SKJ as tenant rather than landlord; "
                "current lease records need a party-role/headlease layer before clean import.",
                "Tenancies",
                headlease_rows,
                "next",
            )
        )
    if arrears_rows or arrears_sheet:
        source_sheet = "Tenancies"
        if arrears_sheet and arrears_rows:
            source_sheet = "Arrears / Tenancies"
        elif arrears_sheet:
            source_sheet = "Arrears"
        feature_candidates.append(
            _feature(
                "arrears_credit_control",
                "Arrears and credit-control queue",
                "Tenancy status and arrears notes can seed an arrears workflow "
                "beyond basic lease status.",
                source_sheet,
                arrears_rows + len(arrears_sheet),
                "next",
            )
        )

    blockers = sum(1 for finding in findings if finding["severity"] == "blocker")
    warnings = sum(1 for finding in findings if finding["severity"] == "warning")
    actions = [
        _action(
            "properties",
            create=property_create,
            match=property_match,
            review=property_review,
        ),
        _action("tenancy_units", create=unit_create, match=unit_match, review=unit_review),
        _action("tenants", create=tenant_create, match=tenant_match),
        _action("leases", create=lease_create, skip=lease_skip, review=lease_review),
        _action("rent_charge_rules", create=rent_rules + outgoings_rules),
        _action(
            "obligations",
            create=date_obligations + bond_security_obligations + bond_insurance_obligations,
        ),
        _action("operational_tasks", create=len(issues) + len(actions_rows)),
    ]
    totals = {
        "sheets": len(sheets),
        "properties": len(properties),
        "tenancies": len(tenancies),
        "bonds": len(bonds),
        "critical_dates": len(dates),
        "vendors": len(vendors),
        "entities": len(entities),
        "active_issues": len(issues),
        "actions": len(actions_rows),
        "blockers": blockers,
        "warnings": warnings,
        "inactive_properties": inactive_properties,
        "vacant_units": vacant_rows,
        "archived_tenancies": archive_rows,
        "headlease_rows": headlease_rows,
    }
    summary = (
        f"Dry run found {property_create} new properties, {unit_create} units, "
        f"{tenant_create} tenants, {lease_create} leases, and "
        f"{date_obligations + bond_security_obligations + bond_insurance_obligations} "
        "obligations to stage."
    )
    return {
        "entity_id": entity_id,
        "filename": filename,
        "sheets": [
            {"name": sheet.name, "rows": len(sheet.rows), "columns": sheet.columns}
            for sheet in sheets.values()
        ],
        "actions": actions,
        "action_items": action_items,
        "findings": findings,
        "feature_candidates": feature_candidates,
        "totals": totals,
        "importable": blockers == 0,
        "summary": summary,
    }


def _confidence_band(confidence: Any) -> str:
    if not isinstance(confidence, (int, float)):
        return "unknown"
    if confidence >= 0.85:
        return "high"
    if confidence >= 0.6:
        return "medium"
    return "low"


def summarize_register_import_plan(action_items: list[dict[str, Any]]) -> dict[str, Any]:
    """Read-only projection bucketing plan rows for operator review.

    Purely additive: derived from a stored plan's action items. It does not
    read or write the database, does not change extraction or decisions, and is
    never on the Apply path.
    """

    by_decision: Counter[str] = Counter()
    by_operation: Counter[str] = Counter()
    by_confidence_band: Counter[str] = Counter()
    blocked_rows = 0
    warning_rows = 0
    ready_to_approve = 0
    needs_attention = 0

    for item in action_items:
        if not isinstance(item, dict):
            continue
        decision = str(item.get("default_decision") or "review")
        operation = str(item.get("operation") or "review")
        source = item.get("source") if isinstance(item.get("source"), dict) else {}
        blockers = item.get("blockers") or []
        warnings = item.get("warnings") or []
        by_decision[decision] += 1
        by_operation[operation] += 1
        by_confidence_band[_confidence_band(source.get("confidence"))] += 1
        if blockers:
            blocked_rows += 1
        if warnings:
            warning_rows += 1
        if blockers or decision == "review":
            needs_attention += 1
        elif decision == "approve":
            ready_to_approve += 1

    return {
        "total_action_items": len([item for item in action_items if isinstance(item, dict)]),
        "by_decision": dict(by_decision),
        "by_operation": dict(by_operation),
        "by_confidence_band": dict(by_confidence_band),
        "blocked_rows": blocked_rows,
        "warning_rows": warning_rows,
        "ready_to_approve": ready_to_approve,
        "needs_attention": needs_attention,
    }


def _date_from_payload(value: Any) -> date | None:
    return _date_value(value)


def _enum_value(enum_type: Any, value: Any, fallback: Any) -> Any:
    try:
        return enum_type(value)
    except (TypeError, ValueError):
        return fallback


def _increment(counter: dict[str, int], key: str, amount: int = 1) -> None:
    counter[key] = counter.get(key, 0) + amount


def _source_from_item(item: dict[str, Any]) -> dict[str, Any]:
    source = item.get("source") if isinstance(item.get("source"), dict) else {}
    return {
        "source": "register_import",
        "action_id": item.get("id"),
        "filename": source.get("filename"),
        "sheet": source.get("sheet"),
        "row": source.get("row"),
        "source_hint": source.get("source_hint"),
        "confidence": source.get("confidence"),
    }


def _metadata_with_register_import(
    metadata: dict[str, Any] | None,
    item: dict[str, Any],
    *,
    changes: list[dict[str, Any]] | None = None,
    payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    next_metadata = dict(metadata or {})
    source = _source_from_item(item)
    history = list(next_metadata.get("register_import_history") or [])
    entry = {
        "action_id": item.get("id"),
        "filename": source.get("filename"),
        "sheet": source.get("sheet"),
        "row": source.get("row"),
        "source_hint": source.get("source_hint"),
        "confidence": source.get("confidence"),
        "changes": _jsonable(changes or item.get("changes") or []),
    }
    if payload:
        entry["payload"] = _jsonable(payload)
    history.append(entry)
    next_metadata.update(
        {
            "source": next_metadata.get("source") or "register_import",
            "last_register_import": entry,
            "register_import_history": history,
        }
    )
    citations = dict(next_metadata.get("source_citations") or {})
    for change in entry["changes"]:
        field = change.get("field") if isinstance(change, dict) else None
        change_source = change.get("source") if isinstance(change, dict) else None
        if field and isinstance(change_source, dict):
            citations[field] = change_source
    if citations:
        next_metadata["source_citations"] = citations
    return next_metadata


def _find_property(
    session: Session,
    entity_id: UUID,
    *,
    code: str | None = None,
    address: str | None = None,
) -> Property | None:
    properties = list(
        session.scalars(
            select(Property).where(Property.entity_id == entity_id, Property.deleted_at.is_(None))
        )
    )
    code_key = _key(code)
    address_key = _key(address)
    for prop in properties:
        if code_key and _key((prop.property_metadata or {}).get("portfolio_code")) == code_key:
            return prop
    for prop in properties:
        if address_key and _key(prop.street_address) == address_key:
            return prop
    return None


def _find_tenant(
    session: Session,
    entity_id: UUID,
    *,
    legal_name: str | None = None,
    abn: str | None = None,
) -> Tenant | None:
    tenants = list(
        session.scalars(
            select(Tenant).where(Tenant.entity_id == entity_id, Tenant.deleted_at.is_(None))
        )
    )
    abn_key = _key(abn)
    name_key = _key(legal_name)
    for tenant in tenants:
        if abn_key and _key(tenant.abn) == abn_key:
            return tenant
    for tenant in tenants:
        if name_key and _key(tenant.legal_name) == name_key:
            return tenant
    return None


def _find_unit(session: Session, property_id: UUID, label: str | None) -> TenancyUnit | None:
    label_key = _key(label)
    if not label_key:
        return None
    units = list(
        session.scalars(
            select(TenancyUnit).where(
                TenancyUnit.property_id == property_id,
                TenancyUnit.deleted_at.is_(None),
            )
        )
    )
    for unit in units:
        if _key(unit.unit_label) == label_key:
            return unit
    return None


def _find_lease_by_ref(session: Session, entity_id: UUID, tenancy_ref: str | None) -> Lease | None:
    ref_key = _key(tenancy_ref)
    if not ref_key:
        return None
    leases = list(
        session.scalars(
            select(Lease)
            .join(TenancyUnit)
            .join(Property)
            .where(
                Property.entity_id == entity_id,
                Property.deleted_at.is_(None),
                TenancyUnit.deleted_at.is_(None),
                Lease.deleted_at.is_(None),
            )
        )
    )
    for lease in leases:
        if _key((lease.lease_metadata or {}).get("portfolio_tenancy_id")) == ref_key:
            return lease
    return None


def _apply_property_item(
    session: Session,
    entity_id: UUID,
    item: dict[str, Any],
) -> dict[str, Any]:
    payload = item.get("payload") if isinstance(item.get("payload"), dict) else {}
    code = _text(payload.get("portfolio_code"))
    address = _text(payload.get("street_address"))
    existing = _find_property(session, entity_id, code=code, address=address)
    operation = item.get("operation")
    if operation == "create":
        if existing is not None:
            return {
                "status": "skipped",
                "message": "Property already exists.",
                "target_table": "property",
                "target_id": existing.id,
            }
        prop = Property(
            entity_id=entity_id,
            name=_text(payload.get("name")) or address or code or "Property to confirm",
            street_address=address or "Address to confirm",
            suburb=_text(payload.get("suburb")) or None,
            state=_text(payload.get("state")) or None,
            postcode=_text(payload.get("postcode")) or None,
            country_code=_text(payload.get("country_code")) or "AU",
            property_type=_enum_value(
                PropertyType,
                payload.get("property_type"),
                PropertyType.other,
            ),
            has_solar_pv=False,
            ownership_structure=_text(payload.get("ownership_structure")) or None,
            owner_legal_name=_text(payload.get("owner_legal_name")) or None,
            invoice_reference=_text(payload.get("invoice_reference")) or None,
            property_metadata=_metadata_with_register_import(
                {
                    "source": "register_import",
                    "portfolio_code": code or None,
                    "source_filename": item["source"]["filename"],
                    "source_sheet": item["source"]["sheet"],
                    "source_row": item["source"].get("row"),
                    "source_hint": item["source"].get("source_hint"),
                    "confidence": item["source"].get("confidence"),
                    "source_notes": payload.get("source_notes"),
                    "source_property_type": payload.get("source_property_type"),
                    "active_tenancies": payload.get("active_tenancies"),
                },
                item,
                payload=payload,
            ),
        )
        session.add(prop)
        session.flush()
        return {
            "status": "applied",
            "message": "Created property.",
            "target_table": "property",
            "target_id": prop.id,
            "created": {"properties": 1},
        }

    if operation == "update":
        if existing is None:
            return {"status": "blocked", "message": "Matched property was not found."}
        applied_changes: list[dict[str, Any]] = []
        for change in item.get("changes") or []:
            field = change.get("field") if isinstance(change, dict) else None
            if field not in {
                "suburb",
                "ownership_structure",
                "owner_legal_name",
                "invoice_reference",
                "property_type",
            }:
                continue
            after = change.get("after")
            current = getattr(existing, field)
            if field == "property_type":
                if current != PropertyType.other:
                    continue
                after = _enum_value(PropertyType, after, current)
            elif current not in (None, ""):
                continue
            setattr(existing, field, after)
            applied_changes.append(
                {**change, "before": _jsonable(current), "after": _jsonable(after)}
            )
        if not applied_changes:
            return {
                "status": "skipped",
                "message": "No blank property fields remained to update.",
                "target_table": "property",
                "target_id": existing.id,
            }
        existing.property_metadata = _metadata_with_register_import(
            {
                **(existing.property_metadata or {}),
                "portfolio_code": code or (existing.property_metadata or {}).get("portfolio_code"),
            },
            item,
            changes=applied_changes,
            payload=payload,
        )
        session.flush()
        return {
            "status": "applied",
            "message": "Updated property.",
            "target_table": "property",
            "target_id": existing.id,
            "updated": {"properties": 1},
        }

    return {"status": "skipped", "message": "Property action was not applyable."}


def _tenant_metadata(payload: dict[str, Any], item: dict[str, Any]) -> dict[str, Any]:
    return _metadata_with_register_import(
        {
            "source": "register_import",
            "source_filename": item["source"]["filename"],
            "source_sheet": item["source"]["sheet"],
            "source_row": item["source"].get("row"),
            "source_hint": item["source"].get("source_hint"),
            "confidence": item["source"].get("confidence"),
            "portfolio_tenancy_id": payload.get("portfolio_tenancy_id"),
        },
        item,
        payload=payload,
    )


def _create_charge_rule(
    session: Session,
    lease: Lease,
    item: dict[str, Any],
    payload: dict[str, Any],
    charge_type: RentChargeType,
    amount_cents: int,
    frequency: RentFrequency,
) -> RentChargeRule:
    rule = RentChargeRule(
        lease_id=lease.id,
        charge_type=charge_type,
        amount_cents=amount_cents,
        frequency=frequency,
        gst_treatment=GstTreatment.taxable,
        start_date=lease.commencement_date,
        end_date=lease.expiry_date,
        next_due_date=lease.commencement_date,
        arrears_or_advance="advance",
        charge_rule_metadata=_metadata_with_register_import(
            {
                "source": "register_import",
                "draft": True,
                "draft_status": "needs_review",
                "source_filename": item["source"]["filename"],
                "source_sheet": item["source"]["sheet"],
                "source_row": item["source"].get("row"),
                "source_hint": item["source"].get("source_hint"),
                "confidence": item["source"].get("confidence"),
                "portfolio_tenancy_id": payload.get("portfolio_tenancy_id"),
                "annual_rent_cents": payload.get("annual_rent_cents"),
            },
            item,
            payload=payload,
        ),
    )
    session.add(rule)
    session.flush()
    return rule


def _apply_tenancy_item(
    session: Session,
    entity_id: UUID,
    item: dict[str, Any],
) -> dict[str, Any]:
    payload = item.get("payload") if isinstance(item.get("payload"), dict) else {}
    tenancy_ref = _text(payload.get("portfolio_tenancy_id"))
    existing_lease = _find_lease_by_ref(session, entity_id, tenancy_ref)
    if existing_lease is not None:
        return {
            "status": "skipped",
            "message": "Lease with this portfolio tenancy ID already exists.",
            "target_table": "lease",
            "target_id": existing_lease.id,
        }

    prop = _find_property(session, entity_id, code=_text(payload.get("property_code")))
    if prop is None:
        return {"status": "blocked", "message": "Property must exist before tenancy apply."}

    created: dict[str, int] = {}
    updated: dict[str, int] = {}
    unit = _find_unit(session, prop.id, _text(payload.get("unit_label")))
    if unit is None:
        unit = TenancyUnit(
            property_id=prop.id,
            unit_label=_text(payload.get("unit_label")) or tenancy_ref or "Unit to confirm",
            sqm=_float_value(payload.get("unit_sqm")),
            unit_metadata=_metadata_with_register_import(
                {
                    "source": "register_import",
                    "source_filename": item["source"]["filename"],
                    "source_sheet": item["source"]["sheet"],
                    "source_row": item["source"].get("row"),
                    "source_hint": item["source"].get("source_hint"),
                    "confidence": item["source"].get("confidence"),
                    "portfolio_tenancy_id": tenancy_ref,
                    "portfolio_property_code": payload.get("property_code"),
                },
                item,
                payload=payload,
            ),
        )
        session.add(unit)
        session.flush()
        _increment(created, "tenancy_units")
    elif unit.sqm is None and payload.get("unit_sqm") is not None:
        before = unit.sqm
        unit.sqm = _float_value(payload.get("unit_sqm"))
        unit.unit_metadata = _metadata_with_register_import(
            unit.unit_metadata,
            item,
            changes=[
                _change(
                    item["source"]["filename"],
                    item["source"]["sheet"],
                    item["source"].get("row"),
                    "sqm",
                    "Size m²",
                    before,
                    unit.sqm,
                )
            ],
            payload=payload,
        )
        _increment(updated, "tenancy_units")

    tenant = _find_tenant(
        session,
        entity_id,
        legal_name=_text(payload.get("tenant_legal_name")),
    )
    if tenant is None:
        tenant = Tenant(
            entity_id=entity_id,
            legal_name=_text(payload.get("tenant_legal_name")) or "Tenant to confirm",
            trading_name=_text(payload.get("tenant_trading_name")) or None,
            contact_name=_text(payload.get("contact_name")) or None,
            tenant_metadata=_tenant_metadata(payload, item),
        )
        session.add(tenant)
        session.flush()
        _increment(created, "tenants")
    else:
        tenant_changes: list[dict[str, Any]] = []
        for field, payload_key, label in (
            ("trading_name", "tenant_trading_name", "Trading name"),
            ("contact_name", "contact_name", "Primary contact"),
        ):
            after = _text(payload.get(payload_key)) or None
            before = getattr(tenant, field)
            if after is not None and before in (None, ""):
                setattr(tenant, field, after)
                tenant_changes.append(
                    _change(
                        item["source"]["filename"],
                        item["source"]["sheet"],
                        item["source"].get("row"),
                        field,
                        label,
                        before,
                        after,
                    )
                )
        if tenant_changes:
            tenant.tenant_metadata = _metadata_with_register_import(
                tenant.tenant_metadata,
                item,
                changes=tenant_changes,
                payload=payload,
            )
            _increment(updated, "tenants")

    rent_frequency = _enum_value(
        RentFrequency,
        payload.get("rent_frequency"),
        RentFrequency.monthly,
    )
    annual_rent_cents = payload.get("annual_rent_cents")
    lease = Lease(
        tenancy_unit_id=unit.id,
        tenant_id=tenant.id,
        status=_enum_value(LeaseStatus, payload.get("lease_status"), LeaseStatus.pending),
        commencement_date=_date_from_payload(payload.get("commencement_date")),
        expiry_date=_date_from_payload(payload.get("expiry_date")),
        annual_rent_cents=annual_rent_cents,
        rent_frequency=rent_frequency,
        outgoings_recoverable=payload.get("outgoings_amount_cents") is not None,
        next_review_date=_date_from_payload(payload.get("next_review_date")),
        option_summary=_text(payload.get("option_summary")) or None,
        security_summary=_text(payload.get("security_summary")) or None,
        notes=_text(payload.get("notes")) or "Created from reviewed register import.",
        lease_metadata=_metadata_with_register_import(
            {
                "source": "register_import",
                "source_filename": item["source"]["filename"],
                "source_sheet": item["source"]["sheet"],
                "source_row": item["source"].get("row"),
                "source_hint": item["source"].get("source_hint"),
                "confidence": item["source"].get("confidence"),
                "portfolio_tenancy_id": tenancy_ref,
                "portfolio_property_code": payload.get("property_code"),
                "review_type": payload.get("review_type"),
                "insurance_status": payload.get("insurance_status"),
                "arrears": payload.get("arrears"),
                "form": payload.get("form"),
            },
            item,
            payload=payload,
        ),
    )
    session.add(lease)
    session.flush()
    _increment(created, "leases")

    if isinstance(annual_rent_cents, int) and annual_rent_cents > 0:
        _create_charge_rule(
            session,
            lease,
            item,
            payload,
            RentChargeType.base_rent,
            _periodic_rent_cents(annual_rent_cents, rent_frequency),
            rent_frequency,
        )
        _increment(created, "rent_charge_rules")
    outgoings_cents = payload.get("outgoings_amount_cents")
    if isinstance(outgoings_cents, int) and outgoings_cents > 0:
        _create_charge_rule(
            session,
            lease,
            item,
            payload,
            RentChargeType.outgoings,
            outgoings_cents,
            rent_frequency,
        )
        _increment(created, "rent_charge_rules")

    return {
        "status": "applied",
        "message": "Created tenancy records.",
        "target_table": "lease",
        "target_id": lease.id,
        "created": created,
        "updated": updated,
    }


def _owner_role(value: Any) -> UserRole | None:
    text = _key(value)
    for role in UserRole:
        if text == role.value:
            return role
    if text in {"accounts", "accounting"}:
        return UserRole.finance
    if text in {"operations", "property"}:
        return UserRole.ops
    return None


def _find_existing_obligation_for_action(
    session: Session,
    entity_id: UUID,
    action_id: str,
) -> Obligation | None:
    obligations = list(
        session.scalars(
            select(Obligation).where(
                Obligation.entity_id == entity_id,
                Obligation.deleted_at.is_(None),
            )
        )
    )
    for obligation in obligations:
        for entry in (obligation.obligation_metadata or {}).get("register_import_history") or []:
            if isinstance(entry, dict) and entry.get("action_id") == action_id:
                return obligation
    return None


def _apply_obligation_item(
    session: Session,
    entity_id: UUID,
    item: dict[str, Any],
) -> dict[str, Any]:
    existing = _find_existing_obligation_for_action(session, entity_id, str(item.get("id")))
    if existing is not None:
        return {
            "status": "skipped",
            "message": "Obligation was already created from this action.",
            "target_table": "obligation",
            "target_id": existing.id,
        }
    payload = item.get("payload") if isinstance(item.get("payload"), dict) else {}
    due_date = _date_from_payload(payload.get("due_date"))
    if due_date is None:
        return {"status": "blocked", "message": "Obligation needs a concrete due date."}
    prop = _find_property(session, entity_id, code=_text(payload.get("property_code")))
    if payload.get("property_code") and prop is None:
        return {"status": "blocked", "message": "Property must exist before obligation apply."}
    lease = _find_lease_by_ref(session, entity_id, _text(payload.get("tenancy_ref")))
    unit_id = lease.tenancy_unit_id if lease is not None else None
    if lease is not None and prop is None:
        unit = session.get(TenancyUnit, lease.tenancy_unit_id)
        if unit is not None:
            prop = session.get(Property, unit.property_id)
    obligation = Obligation(
        entity_id=entity_id,
        property_id=prop.id if prop is not None else None,
        tenancy_unit_id=unit_id,
        lease_id=lease.id if lease is not None else None,
        title=_text(payload.get("title")) or "Workbook obligation",
        category=_enum_value(
            ObligationCategory,
            payload.get("category"),
            ObligationCategory.other,
        ),
        status=ObligationStatus.upcoming,
        due_date=due_date,
        priority=2,
        owner_role=_owner_role(payload.get("owner_role")),
        notes=_text(payload.get("notes")) or None,
        obligation_metadata=_metadata_with_register_import(
            {
                "source": "register_import",
                "source_filename": item["source"]["filename"],
                "source_sheet": item["source"]["sheet"],
                "source_row": item["source"].get("row"),
                "source_hint": item["source"].get("source_hint"),
                "confidence": item["source"].get("confidence"),
                "portfolio_property_code": payload.get("property_code"),
                "portfolio_tenancy_id": payload.get("tenancy_ref"),
            },
            item,
            payload=payload,
        ),
    )
    session.add(obligation)
    session.flush()
    return {
        "status": "applied",
        "message": "Created obligation.",
        "target_table": "obligation",
        "target_id": obligation.id,
        "created": {"obligations": 1},
    }


def apply_register_import_plan(
    *,
    session: Session,
    entity_id: UUID,
    filename: str,
    action_items: list[dict[str, Any]],
    approved_action_ids: list[str],
    ignored_action_ids: list[str] | None = None,
) -> dict[str, Any]:
    """Apply only the reviewed register import actions explicitly approved by a user."""

    approved = set(approved_action_ids)
    ignored = set(ignored_action_ids or [])
    if not approved:
        raise RegisterImportError("Approve at least one register import action before applying.")

    items_by_id = {str(item.get("id")): item for item in action_items}
    missing_ids = sorted(action_id for action_id in approved if action_id not in items_by_id)
    if missing_ids:
        raise RegisterImportError("Approved register import action was not present in the plan.")

    created: dict[str, int] = {}
    updated: dict[str, int] = {}
    results: list[dict[str, Any]] = []
    applied_count = 0
    skipped_count = 0
    blocked_count = 0

    for item in action_items:
        action_id = str(item.get("id"))
        if action_id not in approved:
            continue
        target = str(item.get("target"))
        operation = str(item.get("operation"))
        item_result: dict[str, Any]
        if item.get("blockers"):
            item_result = {"status": "blocked", "message": "; ".join(item["blockers"])}
        elif target == "properties" and operation in {"create", "update"}:
            item_result = _apply_property_item(session, entity_id, item)
        elif target == "tenancies" and operation == "create":
            item_result = _apply_tenancy_item(session, entity_id, item)
        elif target == "obligations" and operation == "create":
            item_result = _apply_obligation_item(session, entity_id, item)
        else:
            item_result = {"status": "skipped", "message": "Action is review-only in v1."}

        for key, value in (item_result.get("created") or {}).items():
            _increment(created, key, value)
        for key, value in (item_result.get("updated") or {}).items():
            _increment(updated, key, value)

        status_value = item_result.get("status")
        if status_value == "applied":
            applied_count += 1
        elif status_value == "blocked":
            blocked_count += 1
        else:
            skipped_count += 1
        results.append(
            {
                "action_id": action_id,
                "target": target,
                "operation": operation,
                "status": status_value,
                "message": item_result.get("message") or "",
                "target_table": item_result.get("target_table"),
                "target_id": item_result.get("target_id"),
                "created": item_result.get("created") or {},
                "updated": item_result.get("updated") or {},
            }
        )

    return {
        "entity_id": entity_id,
        "filename": filename,
        "applied_at": utcnow(),
        "requested": len(approved),
        "applied": applied_count,
        "skipped": skipped_count,
        "blocked": blocked_count,
        "created": created,
        "updated": updated,
        "ignored_action_ids": sorted(ignored),
        "results": results,
    }
