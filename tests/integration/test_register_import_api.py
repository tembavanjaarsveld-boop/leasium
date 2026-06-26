"""Register spreadsheet import dry-run tests."""

from collections.abc import Generator
from io import BytesIO
from uuid import UUID

from apps.api.deps import get_session
from apps.api.main import app
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session
from stewart.core.models import (
    Entity,
    Lease,
    Obligation,
    Property,
    RegisterImportPlan,
    RentChargeRule,
    TenancyUnit,
    Tenant,
)
from stewart.core.settings import get_settings


def _entity_id(session: Session) -> str:
    entity = session.scalar(select(Entity).where(Entity.name == "SKJ Property Pty Ltd"))
    assert entity is not None
    return str(entity.id)


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    properties = workbook.active
    assert properties is not None
    properties.title = "Properties"
    properties.append(
        [
            "Code",
            "Suburb",
            "Address",
            "Owning entity (legal)",
            "Role",
            "Property type",
            "Active tenancies",
            "Status",
            "Notes",
        ]
    )
    properties.append(
        [
            "NL1642",
            "North Lakes",
            "1642 Anzac Avenue",
            "SKJ Property Pty Ltd",
            "Landlord",
            "Commercial multi-tenancy",
            "2",
            "Active",
            "Source row",
        ]
    )
    properties.append(
        [
            "BK62",
            "Brendale",
            "62 Kremzow Road",
            "SJI No 8 Pty Ltd",
            "Tenant (SJI No 8 Pty Ltd)",
            "Brewery main + storage",
            "1",
            "Active",
            "Head lease needs review",
        ]
    )

    tenancies = workbook.create_sheet("Tenancies")
    tenancies.append(
        [
            "Tenancy ID",
            "Property",
            "Unit code",
            "Tenant (legal name)",
            "Trading name",
            "Size m²",
            "Commencement",
            "Expiry",
            "Status",
            "Annual rent",
            "Outgoings",
            "Security",
            "Frequency",
            "Rent per m²",
            "Form",
            "Insurance",
            "Arrears",
            "Review type",
            "Next review",
            "Options",
            "Primary contact",
            "Notes",
        ]
    )
    tenancies.append(
        [
            "NL1642-T001",
            "NL1642",
            "T001",
            "North Lakes Tenant Pty Ltd",
            "Tenant",
            75,
            "2026-05-01",
            "2029-04-30",
            "Active",
            57000,
            10238.88,
            "Bank guarantee",
            "Monthly",
            None,
            "Commercial Lease Agreement",
            "Current",
            "",
            "CPI",
            "2027-05-01",
            "3 years",
            "Alex Tenant",
            "",
        ]
    )
    tenancies.append(
        [
            "BK62-U2",
            "BK62",
            "U2",
            "SKJ Bottle Shop Pty Ltd",
            "Bottle Shop",
            120,
            "2026-01-01",
            "2028-12-31",
            "SKJ-as-tenant",
            32000,
            None,
            "Cash bond",
            "Monthly",
            None,
            "Lease",
            "Current",
            "",
            "3%",
            "2027-01-01",
            "",
            "",
            "Head lease",
        ]
    )
    tenancies.append(
        [
            "",
            "NL1642",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    bonds = workbook.create_sheet("Bonds")
    bonds.append(
        [
            "Tenancy",
            "Property",
            "Tenant",
            "Security type",
            "Amount $AUD",
            "Paid date",
            "Security expiry",
            "Insurance status",
            "Insurance expiry",
            "Notes",
        ]
    )
    bonds.append(
        [
            "NL1642-T001",
            "NL1642",
            "North Lakes Tenant",
            "Bank guarantee",
            39960,
            None,
            "Termination date",
            "Current",
            "2027-07-31",
            "Original held",
        ]
    )

    dates = workbook.create_sheet("Dates")
    dates.append(["Date", "Property", "Tenancy", "Event type", "Description", "Severity", "Owner"])
    dates.append(
        [
            "2027-05-01",
            "NL1642",
            "NL1642-T001",
            "CPI review",
            "Annual rent review",
            "Medium",
            "Finance",
        ]
    )

    vendors = workbook.create_sheet("Vendors")
    vendors.append(
        [
            "Vendor / Counterparty",
            "Category",
            "Scope",
            "Sites / properties",
            "Contact",
            "Status",
            "Notes",
        ]
    )
    vendors.append(["Evolt", "Trades", "Electrical", "All", "TBC", "Active", "Panel"])

    issues = workbook.create_sheet("Active Issues")
    issues.append(["Issue", "Severity", "Owner", "Status", "Notes / next step"])
    issues.append(["Roof leak", "High", "Ops", "Open", "Repair"])

    actions = workbook.create_sheet("Actions")
    actions.append(["Deadline", "Action", "Detail", "Owner"])
    actions.append(["2026-05-20", "Confirm tenant details", "Call tenant", "Ops"])

    charge_rules = workbook.create_sheet("Charge Rules")
    charge_rules.append(
        [
            "Tenancy ID",
            "Charge type",
            "Amount",
            "Frequency",
            "GST treatment",
            "Start date",
            "End date",
            "Next due date",
            "Billing identity",
            "Notes",
            "Source / confidence hint",
        ]
    )
    charge_rules.append(
        [
            "NL1642-T001",
            "Promotion levy",
            120,
            "Monthly",
            "GST",
            "2026-05-01",
            "",
            "2026-06-01",
            "SKJ Property Pty Ltd",
            "Needs mapping",
            "Migration workbook",
        ]
    )

    arrears = workbook.create_sheet("Arrears")
    arrears.append(
        [
            "Tenancy ID",
            "Tenant",
            "Amount overdue",
            "Days overdue",
            "Last contact",
            "Promise to pay",
            "Owner",
            "Status",
            "Notes",
            "Source / confidence hint",
        ]
    )
    arrears.append(
        [
            "NL1642-T001",
            "North Lakes Tenant",
            880,
            14,
            "2026-05-18",
            "2026-05-30",
            "Finance",
            "Open",
            "Call logged",
            "Migration workbook",
        ]
    )

    entities = workbook.create_sheet("Entities")
    entities.append(["Entity (legal name)", "Type", "Role", "Properties", "Notes"])
    entities.append(["SKJ Property Pty Ltd", "Pty Ltd", "Landlord", "NL1642", ""])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_register_import_template_downloads_no_mutation_workbook(
    client: TestClient,
    session: Session,
) -> None:
    plan_count_before = len(list(session.scalars(select(RegisterImportPlan))))
    response = client.get("/api/v1/register-imports/template")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "relby-migration-template.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == [
        "Instructions",
        "Entities",
        "Properties",
        "Tenancies",
        "Charge Rules",
        "Bonds",
        "Dates",
        "Vendors",
        "Arrears",
        "Active Issues",
        "Actions",
    ]
    instructions = workbook["Instructions"]
    instruction_text = " ".join(
        str(cell.value)
        for row in instructions.iter_rows(min_row=1, max_row=8)
        for cell in row
        if cell.value
    )
    assert "Download this template" in instruction_text
    assert "dry-run" in instruction_text
    assert "explicit Apply" in instruction_text
    assert "Nothing changes" in instruction_text

    properties_headers = [cell.value for cell in workbook["Properties"][1]]
    assert properties_headers[:5] == [
        "Code",
        "Suburb",
        "Address",
        "Owning entity (legal)",
        "Role",
    ]
    tenancy_headers = [cell.value for cell in workbook["Tenancies"][1]]
    assert "Unit code" in tenancy_headers
    assert "Tenant (legal name)" in tenancy_headers
    assert "Annual rent" in tenancy_headers
    assert "Source / confidence hint" in tenancy_headers
    assert [cell.value for cell in workbook["Arrears"][1]][:4] == [
        "Tenancy ID",
        "Tenant",
        "Amount overdue",
        "Days overdue",
    ]

    assert len(list(session.scalars(select(RegisterImportPlan)))) == plan_count_before
    assert session.scalar(select(Property)) is None


def test_register_import_template_requires_authenticated_operator_in_clerk_mode(
    session: Session,
) -> None:
    def override_session() -> Generator[Session, None, None]:
        yield session

    app.dependency_overrides[get_session] = override_session
    app.dependency_overrides[get_settings] = lambda: get_settings().model_copy(
        update={
            "auth_mode": "clerk",
            "clerk_secret_key": "sk_test_clerk",
            "clerk_jwks_url": "https://clerk.example/.well-known/jwks.json",
            "clerk_allow_legacy_token_mapping": False,
        }
    )
    try:
        with TestClient(app) as clerk_client:
            response = clerk_client.get("/api/v1/register-imports/template")
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 401
    assert response.json()["detail"] == "Missing Clerk bearer token."


def test_register_import_dry_run_plans_workbook_without_mutation(
    client: TestClient,
    session: Session,
) -> None:
    entity_id = _entity_id(session)
    assert session.scalar(select(Property).where(Property.entity_id == UUID(entity_id))) is None
    response = client.post(
        "/api/v1/register-imports/dry-run",
        data={"entity_id": entity_id},
        files={
            "file": (
                "portfolio.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["plan_id"]
    assert body["entity_id"] == entity_id
    assert body["filename"] == "portfolio.xlsx"
    assert body["totals"]["properties"] == 2
    assert body["totals"]["tenancies"] == 3
    assert body["totals"]["headlease_rows"] == 1
    assert body["totals"]["blockers"] == 1
    assert body["importable"] is False
    assert any(
        action["target"] == "properties" and action["create"] == 2 for action in body["actions"]
    )
    assert any(action["target"] == "leases" and action["create"] == 1 for action in body["actions"])
    assert any(candidate["key"] == "vendor_directory" for candidate in body["feature_candidates"])
    assert any(
        candidate["key"] == "headlease_role_model" for candidate in body["feature_candidates"]
    )
    assert any(
        candidate["key"] == "charge_rule_review" for candidate in body["feature_candidates"]
    )
    arrears_candidate = next(
        candidate
        for candidate in body["feature_candidates"]
        if candidate["key"] == "arrears_credit_control"
    )
    assert arrears_candidate["source_sheet"] == "Arrears"
    assert arrears_candidate["source_count"] == 1
    assert any(
        item["target"] == "tenancies"
        and item["operation"] == "create"
        and item["default_decision"] == "approve"
        for item in body["action_items"]
    )

    unchanged = session.scalar(select(Entity).where(Entity.id == UUID(entity_id)))
    assert unchanged is not None
    assert session.scalar(select(Property).where(Property.entity_id == UUID(entity_id))) is None
    plan = session.get(RegisterImportPlan, UUID(body["plan_id"]))
    assert plan is not None
    assert plan.entity_id == UUID(entity_id)
    assert plan.filename == "portfolio.xlsx"
    assert plan.plan_data["filename"] == "portfolio.xlsx"
    assert plan.plan_data["action_items"][0]["id"]


def test_register_import_get_plan_returns_review_summary_without_mutation(
    client: TestClient,
    session: Session,
) -> None:
    entity_id = _entity_id(session)
    dry_run_response = client.post(
        "/api/v1/register-imports/dry-run",
        data={"entity_id": entity_id},
        files={
            "file": (
                "portfolio.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert dry_run_response.status_code == 200
    dry_run = dry_run_response.json()
    plan_id = dry_run["plan_id"]

    property_count_before = len(
        list(session.scalars(select(Property).where(Property.entity_id == UUID(entity_id))))
    )

    response = client.get(
        f"/api/v1/register-imports/{plan_id}",
        params={"entity_id": entity_id},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["plan_id"] == plan_id
    assert body["entity_id"] == entity_id
    assert body["applied"] is False
    assert body["applied_at"] is None
    # Review summary reflects the stored action items.
    summary = body["review_summary"]
    assert summary["total_action_items"] == len(dry_run["action_items"])
    assert sum(summary["by_decision"].values()) == summary["total_action_items"]
    assert sum(summary["by_operation"].values()) == summary["total_action_items"]
    assert sum(summary["by_confidence_band"].values()) == summary["total_action_items"]
    # Decision and operation buckets agree with the raw action items.
    expected_decisions: dict[str, int] = {}
    expected_operations: dict[str, int] = {}
    expected_blocked = 0
    for item in dry_run["action_items"]:
        expected_decisions[item["default_decision"]] = (
            expected_decisions.get(item["default_decision"], 0) + 1
        )
        expected_operations[item["operation"]] = (
            expected_operations.get(item["operation"], 0) + 1
        )
        if item["blockers"]:
            expected_blocked += 1
    assert summary["by_decision"] == expected_decisions
    assert summary["by_operation"] == expected_operations
    assert summary["blocked_rows"] == expected_blocked
    assert summary["ready_to_approve"] >= 1
    assert summary["needs_attention"] >= summary["blocked_rows"]

    # Regression: reading the plan performs no mutation and does not apply it.
    assert len(
        list(session.scalars(select(Property).where(Property.entity_id == UUID(entity_id))))
    ) == property_count_before
    plan = session.get(RegisterImportPlan, UUID(plan_id))
    assert plan is not None
    assert plan.applied_at is None
    assert "apply_result" not in (plan.plan_data or {})


def test_register_import_get_plan_unknown_id_returns_404(
    client: TestClient,
    session: Session,
) -> None:
    entity_id = _entity_id(session)
    response = client.get(
        "/api/v1/register-imports/00000000-0000-0000-0000-000000000000",
        params={"entity_id": entity_id},
    )
    assert response.status_code == 404
    assert response.json()["detail"] == "Register import plan not found."


def test_register_import_apply_creates_approved_records_with_provenance(
    client: TestClient,
    session: Session,
) -> None:
    entity_id = _entity_id(session)
    dry_run_response = client.post(
        "/api/v1/register-imports/dry-run",
        data={"entity_id": entity_id},
        files={
            "file": (
                "portfolio.xlsx",
                _workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert dry_run_response.status_code == 200
    dry_run = dry_run_response.json()
    approved_action_ids = [
        item["id"]
        for item in dry_run["action_items"]
        if item["default_decision"] == "approve" and not item["blockers"]
    ]
    assert approved_action_ids

    apply_response = client.post(
        "/api/v1/register-imports/apply",
        json={
            "entity_id": entity_id,
            "filename": dry_run["filename"],
            "plan_id": dry_run["plan_id"],
            "action_items": [
                {
                    **dry_run["action_items"][0],
                    "id": "tampered-client-copy",
                }
            ],
            "approved_action_ids": approved_action_ids,
            "ignored_action_ids": [
                item["id"]
                for item in dry_run["action_items"]
                if item["id"] not in approved_action_ids
            ],
        },
    )

    assert apply_response.status_code == 200
    body = apply_response.json()
    assert body["requested"] == len(approved_action_ids)
    assert body["blocked"] == 0
    assert body["created"]["properties"] == 2
    assert body["created"]["tenancy_units"] == 1
    assert body["created"]["tenants"] == 1
    assert body["created"]["leases"] == 1
    assert body["created"]["rent_charge_rules"] == 2
    assert body["created"]["obligations"] == 2
    plan = session.get(RegisterImportPlan, UUID(dry_run["plan_id"]))
    assert plan is not None
    assert plan.applied_at is not None
    assert plan.applied_by_user_id is not None
    assert plan.plan_data["approved_action_ids"] == approved_action_ids
    assert plan.plan_data["apply_result"]["applied"] == body["applied"]

    repeated_apply_response = client.post(
        "/api/v1/register-imports/apply",
        json={
            "entity_id": entity_id,
            "filename": dry_run["filename"],
            "plan_id": dry_run["plan_id"],
            "approved_action_ids": approved_action_ids,
        },
    )
    assert repeated_apply_response.status_code == 409
    assert repeated_apply_response.json()["detail"] == (
        "Register import plan has already been applied."
    )

    properties = list(
        session.scalars(select(Property).where(Property.entity_id == UUID(entity_id)))
    )
    assert len(properties) == 2
    north_lakes = next(
        prop for prop in properties if prop.property_metadata["portfolio_code"] == "NL1642"
    )
    assert north_lakes.property_metadata["source"] == "register_import"
    assert north_lakes.property_metadata["last_register_import"]["filename"] == "portfolio.xlsx"
    assert north_lakes.property_metadata["last_register_import"]["sheet"] == "Properties"
    assert north_lakes.property_metadata["last_register_import"]["row"] == 2
    assert any(
        change["field"] == "street_address"
        and change["before"] is None
        and change["after"] == "1642 Anzac Avenue"
        for change in north_lakes.property_metadata["last_register_import"]["changes"]
    )

    unit = session.scalar(select(TenancyUnit).where(TenancyUnit.unit_label == "T001"))
    assert unit is not None
    tenant = session.scalar(select(Tenant).where(Tenant.legal_name == "North Lakes Tenant Pty Ltd"))
    assert tenant is not None
    lease = session.scalar(select(Lease).where(Lease.tenant_id == tenant.id))
    assert lease is not None
    assert lease.tenancy_unit_id == unit.id
    assert lease.lease_metadata["portfolio_tenancy_id"] == "NL1642-T001"
    assert lease.lease_metadata["last_register_import"]["source_hint"] == "Tenancies row 2"

    charge_rules = list(
        session.scalars(select(RentChargeRule).where(RentChargeRule.lease_id == lease.id))
    )
    assert len(charge_rules) == 2
    assert all(rule.charge_rule_metadata["source"] == "register_import" for rule in charge_rules)

    obligations = list(
        session.scalars(select(Obligation).where(Obligation.entity_id == UUID(entity_id)))
    )
    assert len(obligations) == 2
    assert {obligation.category.value for obligation in obligations} == {
        "insurance",
        "rent_review",
    }
    assert all(
        obligation.obligation_metadata["source_filename"] == "portfolio.xlsx"
        for obligation in obligations
    )
